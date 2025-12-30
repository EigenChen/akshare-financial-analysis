"""
财务分析

从三大报表中提取和计算财务相关指标，包括营收基本数据和费用构成数据
"""

import akshare as ak
import pandas as pd
import os
import time
from datetime import datetime
from typing import Optional, Dict

def get_symbol_name(symbol):
    """
    获取股票名称
    
    参数:
        symbol: 股票代码，如 "600519"
    
    返回:
        股票名称
    """
    try:
        symbol_clean = symbol.replace('.SZ', '').replace('.SH', '')
        stock_list = ak.stock_info_a_code_name()
        stock_info = stock_list[stock_list['code'] == symbol_clean]
        if not stock_info.empty:
            return stock_info.iloc[0]['name']
        return symbol_clean
    except:
        return symbol.replace('.SZ', '').replace('.SH', '')

def get_annual_data(symbol, start_year=2015, end_year=2024):
    """
    获取指定年份范围的年报数据
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        包含利润表、现金流量表数据的字典
    """
    # 确保有交易所后缀
    symbol_with_suffix = symbol
    if '.' not in symbol:
        if symbol.startswith(('000', '001', '002', '300')):
            symbol_with_suffix = symbol + '.SZ'
        elif symbol.startswith(('600', '601', '603', '605', '688')):
            symbol_with_suffix = symbol + '.SH'
        else:
            symbol_with_suffix = symbol + '.SZ'
    
    results = {
        'profit': None,
        'cash_flow': None,
        'balance_sheet': None
    }
    
    try:
        # 获取利润表
        print("正在获取利润表数据...")
        profit = ak.stock_profit_sheet_by_report_em(symbol=symbol_with_suffix)
        if profit is not None and not profit.empty:
            results['profit'] = profit
            print(f"✓ 利润表数据获取成功，共 {len(profit)} 条记录")
        
        # 获取现金流量表
        print("正在获取现金流量表数据...")
        cash_flow = ak.stock_cash_flow_sheet_by_report_em(symbol=symbol_with_suffix)
        if cash_flow is not None and not cash_flow.empty:
            results['cash_flow'] = cash_flow
            print(f"✓ 现金流量表数据获取成功，共 {len(cash_flow)} 条记录")
        
        # 获取资产负债表
        print("正在获取资产负债表数据...")
        balance_sheet = ak.stock_balance_sheet_by_report_em(symbol=symbol_with_suffix)
        if balance_sheet is not None and not balance_sheet.empty:
            results['balance_sheet'] = balance_sheet
            print(f"✓ 资产负债表数据获取成功，共 {len(balance_sheet)} 条记录")
        
    except Exception as e:
        print(f"✗ 获取数据失败: {e}")
        import traceback
        traceback.print_exc()
    
    return results

def extract_year_data(df, year, date_col_name='REPORT_DATE'):
    """
    从数据框中提取指定年份的数据
    
    参数:
        df: 数据框
        year: 年份，如 2024
        date_col_name: 报告期列名
    
    返回:
        该年份的数据行（DataFrame），如果没有则返回None
    """
    if df is None or df.empty:
        return None
    
    # 查找报告期列
    date_col = None
    for col in df.columns:
        if date_col_name in col or '报告期' in col:
            date_col = col
            break
    
    if date_col is None:
        return None
    
    # 筛选指定年份的年报数据（12-31）
    year_str = str(year)
    filtered = df[
        (df[date_col].astype(str).str.contains(year_str, na=False)) &
        (df[date_col].astype(str).str.contains('12-31', na=False))
    ]
    
    if not filtered.empty:
        return filtered.iloc[-1]  # 取最后一条（最新的）
    
    return None

def get_value_from_row(row, column_name, default="-", return_numeric=True):
    """
    从数据行中获取指定列的值，转换为数值（亿元）
    
    参数:
        row: 数据行
        column_name: 列名
        default: 默认值，如果数据缺失返回此值（通常为 "-"）
        return_numeric: 是否返回数值，False时返回字符串（用于缺失数据）
    
    返回:
        数值（亿元），保留2位小数；如果数据缺失，返回 default（通常是 "-"）
    """
    if row is None:
        return default
    
    if column_name not in row.index:
        return default
    
    value = row[column_name]
    try:
        num_value = float(value)
        if pd.isna(num_value):
            return default
        # 转换为亿元
        if return_numeric:
            return round(num_value / 100000000, 2)
        else:
            return str(round(num_value / 100000000, 2))
    except (ValueError, TypeError):
        return default

def calculate_revenue_metrics(symbol, start_year, end_year):
    """
    计算营收基本数据指标
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        包含所有指标数据的DataFrame
    """
    print("=" * 80)
    print(f"开始计算 {symbol} 的营收基本数据（{start_year}-{end_year}）")
    print("=" * 80)
    
    # 获取数据
    data = get_annual_data(symbol, start_year, end_year)
    
    if data['profit'] is None or data['cash_flow'] is None:
        print("✗ 数据获取不完整，无法计算")
        return None
    
    # 准备结果数据
    metrics = {
        '科目': [],
    }
    
    # 初始化年份列
    for year in range(start_year, end_year + 1):
        metrics[str(year)] = []
    
    # 先初始化所有科目名称
    metrics['科目'] = [
        '收入（亿元）',
        '归母净利润（亿元）',
        '净利润率（%）',
        '扣非净利润（亿元）',
        '经营净现金流（亿元）',
        'CAPEX（亿元）',
        '自由现金流（亿元）',
        '扣非/净利润（%）',
        '经营现金流/净利润（%）',
        '金融利润（亿元）',
        '经营利润（亿元）',
        '经营利润/归母利润（%）',
        '金融利润/归母利润（%）'
    ]
    
    # 初始化所有年份的数据列表（默认值为 "-" 表示数据缺失）
    for year in range(start_year, end_year + 1):
        for _ in range(len(metrics['科目'])):
            metrics[str(year)].append("-")
    
    # 逐年计算指标
    for year in range(start_year, end_year + 1):
        print(f"\n处理 {year} 年数据...")
        
        # 获取该年份的数据
        profit_row = extract_year_data(data['profit'], year)
        cash_flow_row = extract_year_data(data['cash_flow'], year)
        
        if profit_row is None:
            print(f"  ⚠ {year} 年利润表数据缺失，使用默认值0")
        if cash_flow_row is None:
            print(f"  ⚠ {year} 年现金流量表数据缺失，使用默认值0")
        
        year_idx = year - start_year
        
        # 检查数据是否可用
        data_available = (profit_row is not None) and (cash_flow_row is not None)
        
        if not data_available:
            print(f"  ⚠ {year} 年数据缺失，该年份所有指标显示为 '-'")
            # 所有指标保持为 "-"（已在初始化时设置）
            continue
        
        # 1. 收入（亿元）- 营业收入
        revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
        if revenue == "-":
            print(f"  ⚠ {year} 年营业收入数据缺失")
            continue  # 如果基础数据缺失，跳过该年份
        metrics[str(year)][0] = revenue
        
        # 2. 归母净利润（亿元）
        parent_net_profit = get_value_from_row(profit_row, 'PARENT_NETPROFIT', "-")
        if parent_net_profit == "-":
            print(f"  ⚠ {year} 年归母净利润数据缺失")
            continue
        metrics[str(year)][1] = parent_net_profit
        
        # 3. 净利润率（%）
        net_profit_margin = round((parent_net_profit / revenue * 100) if revenue != 0 else 0, 2)
        metrics[str(year)][2] = net_profit_margin
        
        # 4. 扣非净利润（亿元）
        deduct_parent_net_profit = get_value_from_row(profit_row, 'DEDUCT_PARENT_NETPROFIT', "-")
        metrics[str(year)][3] = deduct_parent_net_profit
        
        # 5. 经营净现金流（亿元）
        operate_cash_flow = get_value_from_row(cash_flow_row, 'NETCASH_OPERATE', "-")
        metrics[str(year)][4] = operate_cash_flow
        
        # 6. CAPEX（亿元）- 购建固定资产、无形资产和其他长期资产支付的现金
        capex = get_value_from_row(cash_flow_row, 'CONSTRUCT_LONG_ASSET', "-")
        metrics[str(year)][5] = capex
        
        # 7. 自由现金流（亿元）= 经营净现金流 - CAPEX
        if operate_cash_flow != "-" and capex != "-":
            free_cash_flow = round(operate_cash_flow - capex, 2)
            metrics[str(year)][6] = free_cash_flow
        else:
            metrics[str(year)][6] = "-"
        
        # 8. 扣非/净利润（%）
        if deduct_parent_net_profit != "-" and parent_net_profit != "-" and parent_net_profit != 0:
            deduct_net_ratio = round((deduct_parent_net_profit / parent_net_profit * 100), 2)
            metrics[str(year)][7] = deduct_net_ratio
        else:
            metrics[str(year)][7] = "-"
        
        # 9. 经营现金流/净利润（%）
        if operate_cash_flow != "-" and parent_net_profit != "-" and parent_net_profit != 0:
            cash_profit_ratio = round((operate_cash_flow / parent_net_profit * 100), 2)
            metrics[str(year)][8] = cash_profit_ratio
        else:
            metrics[str(year)][8] = "-"
        
        # 10. 金融利润（亿元）= 公允价值变动收益 + 投资收益
        # 注意：如果科目不存在或为空，可能返回0或"-"，需要统一处理
        fairvalue_change = get_value_from_row(profit_row, 'FAIRVALUE_CHANGE_INCOME', 0)
        invest_income = get_value_from_row(profit_row, 'INVEST_INCOME', 0)
        # 如果值为 "-" 或 None，视为0；否则计算
        if fairvalue_change == "-" or fairvalue_change is None:
            fairvalue_change = 0
        if invest_income == "-" or invest_income is None:
            invest_income = 0
        financial_profit = round(fairvalue_change + invest_income, 2)
        metrics[str(year)][9] = financial_profit
        
        # 11. 经营利润（亿元）= 归母净利润 - 金融利润
        if parent_net_profit != "-" and metrics[str(year)][9] != "-":
            operate_profit = round(parent_net_profit - metrics[str(year)][9], 2)
            metrics[str(year)][10] = operate_profit
        else:
            metrics[str(year)][10] = "-"
        
        # 12. 经营利润/归母利润（%）
        if metrics[str(year)][10] != "-" and parent_net_profit != "-" and parent_net_profit != 0:
            operate_profit_ratio = round((metrics[str(year)][10] / parent_net_profit * 100), 2)
            metrics[str(year)][11] = operate_profit_ratio
        else:
            metrics[str(year)][11] = "-"
        
        # 13. 金融利润（%）- 金融利润占归母净利润的比例
        if metrics[str(year)][9] != "-" and parent_net_profit != "-" and parent_net_profit != 0:
            financial_profit_ratio = round((metrics[str(year)][9] / parent_net_profit * 100), 2)
            metrics[str(year)][12] = financial_profit_ratio
        else:
            metrics[str(year)][12] = "-"
        
        print(f"  ✓ {year} 年数据计算完成")
    
    # 创建DataFrame
    result_df = pd.DataFrame(metrics)
    
    return result_df

def calculate_expense_metrics(symbol, start_year, end_year):
    """
    计算费用构成指标
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        包含所有费用指标数据的DataFrame
    """
    print("\n" + "=" * 80)
    print(f"开始计算 {symbol} 的费用构成数据（{start_year}-{end_year}）")
    print("=" * 80)
    
    # 获取数据
    data = get_annual_data(symbol, start_year, end_year)
    
    if data['profit'] is None:
        print("✗ 利润表数据获取不完整，无法计算")
        return None
    
    # 准备结果数据
    metrics = {
        '科目': [],
    }
    
    # 初始化年份列
    for year in range(start_year, end_year + 1):
        metrics[str(year)] = []
    
    # 先初始化所有科目名称
    metrics['科目'] = [
        '毛利率（%）',
        '净利率（%）',
        '研发费用（亿元）',
        '管理费用（亿元）',
        '管理研发费用（亿元）',
        '销售费用（亿元）',
        '财务费用（亿元）',
        '期间费用合计（亿元）',
        '销售费用率（%）',
        '研发费用率（%）',
        '管理费用率（%）',
        '管理研发费用率（%）',
        '期间费用率（%）',
        '毛利率-净利润率（%）'
    ]
    
    # 初始化所有年份的数据列表（默认值为 "-" 表示数据缺失）
    for year in range(start_year, end_year + 1):
        for _ in range(len(metrics['科目'])):
            metrics[str(year)].append("-")
    
    # 逐年计算指标
    for year in range(start_year, end_year + 1):
        print(f"\n处理 {year} 年数据...")
        
        # 获取该年份的数据
        profit_row = extract_year_data(data['profit'], year)
        
        if profit_row is None:
            print(f"  ⚠ {year} 年利润表数据缺失，该年份所有指标显示为 '-'")
            continue
        
        # 1. 收入（亿元）- 营业收入（用于计算各种比率）
        revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
        if revenue == "-":
            print(f"  ⚠ {year} 年营业收入数据缺失")
            continue
        
        # 2. 营业成本（亿元）- 用于计算毛利率
        operate_cost = get_value_from_row(profit_row, 'OPERATE_COST', "-")
        
        # 3. 归母净利润（亿元）- 用于计算净利率
        parent_net_profit = get_value_from_row(profit_row, 'PARENT_NETPROFIT', "-")
        if parent_net_profit == "-":
            print(f"  ⚠ {year} 年归母净利润数据缺失")
            continue
        
        # 1. 毛利率（%）= (营业收入 - 营业成本) / 营业收入 * 100
        if revenue != "-" and operate_cost != "-" and revenue != 0:
            gross_margin = round(((revenue - operate_cost) / revenue * 100), 2)
            metrics[str(year)][0] = gross_margin
        else:
            metrics[str(year)][0] = "-"
        
        # 2. 净利率（%）= 归母净利润 / 营业收入 * 100
        if revenue != "-" and parent_net_profit != "-" and revenue != 0:
            net_margin = round((parent_net_profit / revenue * 100), 2)
            metrics[str(year)][1] = net_margin
        else:
            metrics[str(year)][1] = "-"
        
        # 3. 研发费用（亿元）
        research_expense = get_value_from_row(profit_row, 'RESEARCH_EXPENSE', "-")
        metrics[str(year)][2] = research_expense
        
        # 4. 管理费用（亿元）
        manage_expense = get_value_from_row(profit_row, 'MANAGE_EXPENSE', "-")
        metrics[str(year)][3] = manage_expense
        
        # 5. 管理研发费用（亿元）= 管理费用 + 研发费用
        if manage_expense != "-" and research_expense != "-":
            manage_research_expense = round(manage_expense + research_expense, 2)
            metrics[str(year)][4] = manage_research_expense
        elif manage_expense != "-" and research_expense == "-":
            # 如果研发费用缺失，只使用管理费用
            metrics[str(year)][4] = manage_expense
        elif manage_expense == "-" and research_expense != "-":
            # 如果管理费用缺失，只使用研发费用
            metrics[str(year)][4] = research_expense
        else:
            metrics[str(year)][4] = "-"
        
        # 6. 销售费用（亿元）
        sale_expense = get_value_from_row(profit_row, 'SALE_EXPENSE', "-")
        metrics[str(year)][5] = sale_expense
        
        # 7. 财务费用（亿元）
        finance_expense = get_value_from_row(profit_row, 'FINANCE_EXPENSE', "-")
        metrics[str(year)][6] = finance_expense
        
        # 8. 期间费用合计（亿元）= 销售费用 + 研发费用 + 管理费用 + 财务费用
        period_expenses = [sale_expense, research_expense, manage_expense, finance_expense]
        # 检查是否有缺失值
        if all(exp != "-" for exp in period_expenses):
            period_expense_total = round(sum(period_expenses), 2)
            metrics[str(year)][7] = period_expense_total
        else:
            # 如果有缺失值，尝试计算非缺失值的和
            valid_expenses = [exp for exp in period_expenses if exp != "-"]
            if valid_expenses:
                period_expense_total = round(sum(valid_expenses), 2)
                metrics[str(year)][7] = period_expense_total
            else:
                metrics[str(year)][7] = "-"
        
        # 9. 销售费用率（%）= 销售费用 / 营业收入 * 100
        if sale_expense != "-" and revenue != "-" and revenue != 0:
            sale_expense_ratio = round((sale_expense / revenue * 100), 2)
            metrics[str(year)][8] = sale_expense_ratio
        else:
            metrics[str(year)][8] = "-"
        
        # 10. 研发费用率（%）= 研发费用 / 营业收入 * 100
        if research_expense != "-" and revenue != "-" and revenue != 0:
            research_expense_ratio = round((research_expense / revenue * 100), 2)
            metrics[str(year)][9] = research_expense_ratio
        else:
            metrics[str(year)][9] = "-"
        
        # 11. 管理费用率（%）= 管理费用 / 营业收入 * 100
        if manage_expense != "-" and revenue != "-" and revenue != 0:
            manage_expense_ratio = round((manage_expense / revenue * 100), 2)
            metrics[str(year)][10] = manage_expense_ratio
        else:
            metrics[str(year)][10] = "-"
        
        # 12. 管理研发费用率（%）= 管理研发费用 / 营业收入 * 100
        if metrics[str(year)][4] != "-" and revenue != "-" and revenue != 0:
            manage_research_ratio = round((metrics[str(year)][4] / revenue * 100), 2)
            metrics[str(year)][11] = manage_research_ratio
        else:
            metrics[str(year)][11] = "-"
        
        # 13. 期间费用率（%）= 期间费用合计 / 营业收入 * 100
        if metrics[str(year)][7] != "-" and revenue != "-" and revenue != 0:
            period_expense_ratio = round((metrics[str(year)][7] / revenue * 100), 2)
            metrics[str(year)][12] = period_expense_ratio
        else:
            metrics[str(year)][12] = "-"
        
        # 14. 毛利率-净利润率（%）= 毛利率 - 净利润率
        if metrics[str(year)][0] != "-" and metrics[str(year)][1] != "-":
            gross_net_diff = round(metrics[str(year)][0] - metrics[str(year)][1], 2)
            metrics[str(year)][13] = gross_net_diff
        else:
            metrics[str(year)][13] = "-"
        
        print(f"  ✓ {year} 年数据计算完成")
    
    # 创建DataFrame
    result_df = pd.DataFrame(metrics)
    
    return result_df

def calculate_growth_metrics(symbol, start_year, end_year):
    """
    计算增长率指标
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        包含所有增长率指标数据的DataFrame
    """
    print("\n" + "=" * 80)
    print(f"开始计算 {symbol} 的增长率数据（{start_year}-{end_year}）")
    print("=" * 80)
    
    # 获取数据
    data = get_annual_data(symbol, start_year, end_year)
    
    if data['profit'] is None:
        print("✗ 利润表数据获取不完整，无法计算")
        return None
    
    # 准备结果数据
    metrics = {
        '科目': [],
    }
    
    # 初始化年份列
    for year in range(start_year, end_year + 1):
        metrics[str(year)] = []
    
    # 添加复合增长率列
    metrics['最近5年'] = []
    metrics['最近3年'] = []
    
    # 初始化科目名称（只有两行）
    metrics['科目'] = [
        '收入增长率（%）',
        '归母净利润增长率（%）'
    ]
    
    # 初始化所有年份的数据列表（默认值为 "-" 表示数据缺失）
    for year in range(start_year, end_year + 1):
        metrics[str(year)].append("-")  # 收入增长率
        metrics[str(year)].append("-")  # 归母净利润增长率
    
    # 初始化复合增长率列
    metrics['最近5年'].append("-")  # 收入增长率
    metrics['最近5年'].append("-")  # 归母净利润增长率
    metrics['最近3年'].append("-")  # 收入增长率
    metrics['最近3年'].append("-")  # 归母净利润增长率
    
    # 存储每年的收入和归母净利润值（用于计算增长率）
    revenue_values = {}
    profit_values = {}
    
    # 先获取所有年份的原始数据
    for year in range(start_year, end_year + 1):
        profit_row = extract_year_data(data['profit'], year)
        
        if profit_row is None:
            print(f"  ⚠ {year} 年利润表数据缺失")
            continue
        
        # 获取收入
        revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
        if revenue != "-":
            revenue_values[year] = revenue
        
        # 获取归母净利润
        parent_net_profit = get_value_from_row(profit_row, 'PARENT_NETPROFIT', "-")
        if parent_net_profit != "-":
            profit_values[year] = parent_net_profit
    
    # 计算各年份的增长率
    for year in range(start_year, end_year + 1):
        print(f"\n处理 {year} 年数据...")
        
        # 第一年是基准，不需要计算增长率
        if year == start_year:
            print(f"  {year} 年为基准年份，增长率显示为 '-'")
            continue
        
        # 计算收入增长率
        if year in revenue_values and (year - 1) in revenue_values:
            prev_revenue = revenue_values[year - 1]
            curr_revenue = revenue_values[year]
            if prev_revenue != 0:
                revenue_growth = round(((curr_revenue - prev_revenue) / prev_revenue * 100), 2)
                metrics[str(year)][0] = revenue_growth
                print(f"  ✓ {year} 年收入增长率: {revenue_growth}%")
            else:
                print(f"  ⚠ {year - 1} 年收入为0，无法计算增长率")
        else:
            print(f"  ⚠ {year} 年收入数据缺失，无法计算增长率")
        
        # 计算归母净利润增长率
        if year in profit_values and (year - 1) in profit_values:
            prev_profit = profit_values[year - 1]
            curr_profit = profit_values[year]
            if prev_profit != 0:
                profit_growth = round(((curr_profit - prev_profit) / prev_profit * 100), 2)
                metrics[str(year)][1] = profit_growth
                print(f"  ✓ {year} 年归母净利润增长率: {profit_growth}%")
            else:
                print(f"  ⚠ {year - 1} 年归母净利润为0，无法计算增长率")
        else:
            print(f"  ⚠ {year} 年归母净利润数据缺失，无法计算增长率")
    
    # 计算复合增长率
    print(f"\n计算复合增长率...")
    
    # 计算最近5年复合增长率
    # 需要 end_year 和 end_year - 5 年的数据
    if end_year in revenue_values and (end_year - 5) in revenue_values:
        start_revenue = revenue_values[end_year - 5]
        end_revenue = revenue_values[end_year]
        if start_revenue > 0 and end_revenue > 0:
            ratio = end_revenue / start_revenue
            if ratio > 0:
                try:
                    # (第N年/第N-5年)^(1/5) - 1
                    cagr_5_revenue = (pow(ratio, 1/5) - 1) * 100
                    # 检查是否是复数
                    if isinstance(cagr_5_revenue, complex):
                        print(f"  ⚠ 收入最近5年复合增长率计算结果为复数，无法计算")
                    else:
                        metrics['最近5年'][0] = round(cagr_5_revenue, 2)
                        print(f"  ✓ 收入最近5年复合增长率: {metrics['最近5年'][0]}%")
                except (ValueError, TypeError):
                    print(f"  ⚠ 收入最近5年复合增长率计算失败")
            else:
                print(f"  ⚠ 收入比率为负数，无法计算最近5年复合增长率")
        else:
            print(f"  ⚠ {end_year - 5} 年或 {end_year} 年收入为0或负数，无法计算最近5年复合增长率")
    else:
        print(f"  ⚠ 缺少必要数据，无法计算收入最近5年复合增长率")
    
    if end_year in profit_values and (end_year - 5) in profit_values:
        start_profit = profit_values[end_year - 5]
        end_profit = profit_values[end_year]
        if start_profit > 0 and end_profit > 0:
            ratio = end_profit / start_profit
            if ratio > 0:
                try:
                    cagr_5_profit = (pow(ratio, 1/5) - 1) * 100
                    # 检查是否是复数
                    if isinstance(cagr_5_profit, complex):
                        print(f"  ⚠ 归母净利润最近5年复合增长率计算结果为复数，无法计算")
                    else:
                        metrics['最近5年'][1] = round(cagr_5_profit, 2)
                        print(f"  ✓ 归母净利润最近5年复合增长率: {metrics['最近5年'][1]}%")
                except (ValueError, TypeError):
                    print(f"  ⚠ 归母净利润最近5年复合增长率计算失败")
            else:
                print(f"  ⚠ 归母净利润比率为负数，无法计算最近5年复合增长率")
        else:
            print(f"  ⚠ {end_year - 5} 年或 {end_year} 年归母净利润为0或负数，无法计算最近5年复合增长率")
    else:
        print(f"  ⚠ 缺少必要数据，无法计算归母净利润最近5年复合增长率")
    
    # 计算最近3年复合增长率
    # 需要 end_year 和 end_year - 3 年的数据
    if end_year in revenue_values and (end_year - 3) in revenue_values:
        start_revenue = revenue_values[end_year - 3]
        end_revenue = revenue_values[end_year]
        if start_revenue > 0 and end_revenue > 0:
            ratio = end_revenue / start_revenue
            if ratio > 0:
                try:
                    # (第N年/第N-3年)^(1/3) - 1
                    cagr_3_revenue = (pow(ratio, 1/3) - 1) * 100
                    # 检查是否是复数
                    if isinstance(cagr_3_revenue, complex):
                        print(f"  ⚠ 收入最近3年复合增长率计算结果为复数，无法计算")
                    else:
                        metrics['最近3年'][0] = round(cagr_3_revenue, 2)
                        print(f"  ✓ 收入最近3年复合增长率: {metrics['最近3年'][0]}%")
                except (ValueError, TypeError):
                    print(f"  ⚠ 收入最近3年复合增长率计算失败")
            else:
                print(f"  ⚠ 收入比率为负数，无法计算最近3年复合增长率")
        else:
            print(f"  ⚠ {end_year - 3} 年或 {end_year} 年收入为0或负数，无法计算最近3年复合增长率")
    else:
        print(f"  ⚠ 缺少必要数据，无法计算收入最近3年复合增长率")
    
    if end_year in profit_values and (end_year - 3) in profit_values:
        start_profit = profit_values[end_year - 3]
        end_profit = profit_values[end_year]
        if start_profit > 0 and end_profit > 0:
            ratio = end_profit / start_profit
            if ratio > 0:
                try:
                    cagr_3_profit = (pow(ratio, 1/3) - 1) * 100
                    # 检查是否是复数
                    if isinstance(cagr_3_profit, complex):
                        print(f"  ⚠ 归母净利润最近3年复合增长率计算结果为复数，无法计算")
                    else:
                        metrics['最近3年'][1] = round(cagr_3_profit, 2)
                        print(f"  ✓ 归母净利润最近3年复合增长率: {metrics['最近3年'][1]}%")
                except (ValueError, TypeError):
                    print(f"  ⚠ 归母净利润最近3年复合增长率计算失败")
            else:
                print(f"  ⚠ 归母净利润比率为负数，无法计算最近3年复合增长率")
        else:
            print(f"  ⚠ {end_year - 3} 年或 {end_year} 年归母净利润为0或负数，无法计算最近3年复合增长率")
    else:
        print(f"  ⚠ 缺少必要数据，无法计算归母净利润最近3年复合增长率")
    
    # 创建DataFrame，并确保列顺序正确
    # 列顺序：科目、各年份列、最近5年、最近3年
    columns_order = ['科目'] + [str(year) for year in range(start_year, end_year + 1)] + ['最近5年', '最近3年']
    result_df = pd.DataFrame(metrics)
    result_df = result_df[columns_order]
    
    return result_df

def calculate_balance_sheet_metrics(symbol, start_year, end_year):
    """
    计算资产负债指标
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        包含所有资产负债指标数据的DataFrame
    """
    print("\n" + "=" * 80)
    print(f"开始计算 {symbol} 的资产负债数据（{start_year}-{end_year}）")
    print("=" * 80)
    
    # 获取数据
    data = get_annual_data(symbol, start_year, end_year)
    
    if data['balance_sheet'] is None:
        print("✗ 资产负债表数据获取不完整，无法计算")
        return None
    
    # 准备结果数据
    metrics = {
        '科目': [],
    }
    
    # 初始化年份列
    for year in range(start_year, end_year + 1):
        metrics[str(year)] = []
    
    # 初始化所有科目名称
    metrics['科目'] = [
        '流动资产（亿元）',
        '现金（亿元）',
        '存货（亿元）',
        '非流动资产（亿元）',
        '总资产（亿元）',
        '归母净资产（亿元）',
        '狭义无息债务（亿元）',
        '广义无息债务（亿元）',
        '有息债务（亿元）',
        '狭义无息债务/收入（%）',
        '狭义无息债务/总资产（%）',
        '广义无息债务/收入（%）',
        '广义无息债务/总资产（%）',
        '有息债务/总资产（%）',
        '资产负债率（%）'
    ]
    
    # 初始化所有年份的数据列表（默认值为 "-" 表示数据缺失）
    for year in range(start_year, end_year + 1):
        for _ in range(len(metrics['科目'])):
            metrics[str(year)].append("-")
    
    # 逐年计算指标
    for year in range(start_year, end_year + 1):
        print(f"\n处理 {year} 年数据...")
        
        # 获取该年份的数据
        balance_row = extract_year_data(data['balance_sheet'], year)
        profit_row = extract_year_data(data['profit'], year) if data['profit'] is not None else None
        
        if balance_row is None:
            print(f"  ⚠ {year} 年资产负债表数据缺失，该年份所有指标显示为 '-'")
            continue
        
        # 1. 流动资产（亿元）
        current_assets = get_value_from_row(balance_row, 'TOTAL_CURRENT_ASSETS', "-")
        metrics[str(year)][0] = current_assets
        
        # 2. 现金（亿元）- 货币资金
        cash = get_value_from_row(balance_row, 'MONETARYFUNDS', "-")
        metrics[str(year)][1] = cash
        
        # 3. 存货（亿元）
        inventory = get_value_from_row(balance_row, 'INVENTORY', "-")
        metrics[str(year)][2] = inventory
        
        # 4. 非流动资产（亿元）
        non_current_assets = get_value_from_row(balance_row, 'TOTAL_NONCURRENT_ASSETS', "-")
        metrics[str(year)][3] = non_current_assets
        
        # 5. 总资产（亿元）
        total_assets = get_value_from_row(balance_row, 'TOTAL_ASSETS', "-")
        metrics[str(year)][4] = total_assets
        
        # 6. 归母净资产（亿元）- 归属于母公司所有者权益合计
        parent_equity = get_value_from_row(balance_row, 'TOTAL_PARENT_EQUITY', "-")
        metrics[str(year)][5] = parent_equity
        
        # 7. 狭义无息债务（亿元）= 应付账款 + 预收账款 + 合同负债
        accounts_payable = get_value_from_row(balance_row, 'ACCOUNTS_PAYABLE', 0)
        advance_receipts = get_value_from_row(balance_row, 'ADVANCE_RECEIVABLES', 0)
        contract_liabilities = get_value_from_row(balance_row, 'CONTRACT_LIAB', 0)
        
        # 处理缺失值
        if accounts_payable == "-":
            accounts_payable = 0
        if advance_receipts == "-":
            advance_receipts = 0
        if contract_liabilities == "-":
            contract_liabilities = 0
        
        narrow_interest_free_debt = round(accounts_payable + advance_receipts + contract_liabilities, 2)
        metrics[str(year)][6] = narrow_interest_free_debt
        
        # 8. 广义无息债务（亿元）= 应付账款 + 应付票据 + 预收账款 + 合同负债
        note_payable = get_value_from_row(balance_row, 'NOTE_PAYABLE', 0)
        if note_payable == "-":
            note_payable = 0
        
        broad_interest_free_debt = round(accounts_payable + note_payable + advance_receipts + contract_liabilities, 2)
        metrics[str(year)][7] = broad_interest_free_debt
        
        # 9. 有息债务（亿元）- 短期借款 + 长期借款 + 应付债券等
        short_term_loan = get_value_from_row(balance_row, 'SHORT_LOAN', 0)
        long_term_loan = get_value_from_row(balance_row, 'LONG_LOAN', 0)
        bonds_payable = get_value_from_row(balance_row, 'BONDS_PAYABLE', 0)
        
        # 处理缺失值
        if short_term_loan == "-":
            short_term_loan = 0
        if long_term_loan == "-":
            long_term_loan = 0
        if bonds_payable == "-":
            bonds_payable = 0
        
        interest_bearing_debt = round(short_term_loan + long_term_loan + bonds_payable, 2)
        metrics[str(year)][8] = interest_bearing_debt
        
        # 10. 狭义无息债务/收入（%）
        if profit_row is not None:
            revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
            if narrow_interest_free_debt != "-" and revenue != "-" and revenue != 0:
                narrow_interest_free_debt_revenue_ratio = round((narrow_interest_free_debt / revenue * 100), 2)
                metrics[str(year)][9] = narrow_interest_free_debt_revenue_ratio
            else:
                metrics[str(year)][9] = "-"
        else:
            metrics[str(year)][9] = "-"
        
        # 11. 狭义无息债务/总资产（%）
        if narrow_interest_free_debt != "-" and total_assets != "-" and total_assets != 0:
            narrow_interest_free_debt_assets_ratio = round((narrow_interest_free_debt / total_assets * 100), 2)
            metrics[str(year)][10] = narrow_interest_free_debt_assets_ratio
        else:
            metrics[str(year)][10] = "-"
        
        # 12. 广义无息债务/收入（%）
        if profit_row is not None:
            revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
            if broad_interest_free_debt != "-" and revenue != "-" and revenue != 0:
                broad_interest_free_debt_revenue_ratio = round((broad_interest_free_debt / revenue * 100), 2)
                metrics[str(year)][11] = broad_interest_free_debt_revenue_ratio
            else:
                metrics[str(year)][11] = "-"
        else:
            metrics[str(year)][11] = "-"
        
        # 13. 广义无息债务/总资产（%）
        if broad_interest_free_debt != "-" and total_assets != "-" and total_assets != 0:
            broad_interest_free_debt_assets_ratio = round((broad_interest_free_debt / total_assets * 100), 2)
            metrics[str(year)][12] = broad_interest_free_debt_assets_ratio
        else:
            metrics[str(year)][12] = "-"
        
        # 14. 有息债务/总资产（%）
        if interest_bearing_debt != "-" and total_assets != "-" and total_assets != 0:
            interest_bearing_debt_assets_ratio = round((interest_bearing_debt / total_assets * 100), 2)
            metrics[str(year)][13] = interest_bearing_debt_assets_ratio
        else:
            metrics[str(year)][13] = "-"
        
        # 15. 资产负债率（%）= 总负债 / 总资产 * 100
        total_liabilities = get_value_from_row(balance_row, 'TOTAL_LIABILITIES', "-")
        if total_liabilities != "-" and total_assets != "-" and total_assets != 0:
            asset_liability_ratio = round((total_liabilities / total_assets * 100), 2)
            metrics[str(year)][14] = asset_liability_ratio
        else:
            metrics[str(year)][14] = "-"
        
        print(f"  ✓ {year} 年数据计算完成")
    
    # 创建DataFrame
    result_df = pd.DataFrame(metrics)
    
    return result_df

def calculate_wc_metrics(symbol, start_year, end_year):
    """
    计算WC（营运资金）相关指标
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        包含所有WC指标数据的DataFrame
    """
    print("\n" + "=" * 80)
    print(f"开始计算 {symbol} 的WC分析数据（{start_year}-{end_year}）")
    print("=" * 80)
    
    # 获取数据
    data = get_annual_data(symbol, start_year, end_year)
    
    if data['balance_sheet'] is None or data['profit'] is None:
        print("✗ 资产负债表或利润表数据获取不完整，无法计算")
        return None
    
    # 准备结果数据
    metrics = {
        '科目': [],
    }
    
    # 初始化年份列
    for year in range(start_year, end_year + 1):
        metrics[str(year)] = []
    
    # 初始化所有科目名称
    metrics['科目'] = [
        '1元收入需要的WC（元）',
        'WC（亿元）',
        '应收（亿元）',
        '预付（亿元）',
        '存货（亿元）',
        '合同资产（亿元）',
        '应付（亿元）',
        '预收（亿元）',
        '合同负债（亿元）',
        '应收占收入比重（%）',
        '预付占收入比重（%）',
        '存货占收入比重（%）',
        '应付占收入比重（%）',
        '预收占收入比重（%）',
        '合同负债占收入比重（%）',
        '新增WC（亿元）'
    ]
    
    # 初始化所有年份的数据列表（默认值为 "-" 表示数据缺失）
    for year in range(start_year, end_year + 1):
        for _ in range(len(metrics['科目'])):
            metrics[str(year)].append("-")
    
    # 存储每年的WC值（用于计算新增WC）
    wc_values = {}
    
    # 逐年计算指标
    for year in range(start_year, end_year + 1):
        print(f"\n处理 {year} 年数据...")
        
        # 获取该年份的数据
        balance_row = extract_year_data(data['balance_sheet'], year)
        profit_row = extract_year_data(data['profit'], year)
        
        if balance_row is None or profit_row is None:
            print(f"  ⚠ {year} 年数据缺失，该年份所有指标显示为 '-'")
            continue
        
        # 获取收入
        revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
        if revenue == "-":
            print(f"  ⚠ {year} 年营业收入数据缺失")
            continue
        
        # 1. 应收（亿元）- 应收账款
        accounts_receivable = get_value_from_row(balance_row, 'ACCOUNTS_RECE', "-")
        metrics[str(year)][2] = accounts_receivable
        
        # 2. 预付（亿元）- 预付账款
        prepayment = get_value_from_row(balance_row, 'PREPAYMENT', "-")
        metrics[str(year)][3] = prepayment
        
        # 3. 存货（亿元）
        inventory = get_value_from_row(balance_row, 'INVENTORY', "-")
        metrics[str(year)][4] = inventory
        
        # 4. 合同资产（亿元）
        contract_asset = get_value_from_row(balance_row, 'CONTRACT_ASSET', "-")
        metrics[str(year)][5] = contract_asset
        
        # 5. 应付（亿元）- 应付账款
        accounts_payable = get_value_from_row(balance_row, 'ACCOUNTS_PAYABLE', "-")
        metrics[str(year)][6] = accounts_payable
        
        # 6. 预收（亿元）- 预收账款
        advance_receipts = get_value_from_row(balance_row, 'ADVANCE_RECEIVABLES', "-")
        metrics[str(year)][7] = advance_receipts
        
        # 7. 合同负债（亿元）
        contract_liabilities = get_value_from_row(balance_row, 'CONTRACT_LIAB', "-")
        metrics[str(year)][8] = contract_liabilities
        
        # 处理缺失值，用于计算WC
        accounts_receivable_val = accounts_receivable if accounts_receivable != "-" else 0
        prepayment_val = prepayment if prepayment != "-" else 0
        inventory_val = inventory if inventory != "-" else 0
        contract_asset_val = contract_asset if contract_asset != "-" else 0
        accounts_payable_val = accounts_payable if accounts_payable != "-" else 0
        advance_receipts_val = advance_receipts if advance_receipts != "-" else 0
        contract_liabilities_val = contract_liabilities if contract_liabilities != "-" else 0
        
        # 8. WC（亿元）= (应收账款 + 预付账款 + 存货 + 合同资产) - (应付账款 + 预收账款 + 合同负债)
        wc = round((accounts_receivable_val + prepayment_val + inventory_val + contract_asset_val) - 
                   (accounts_payable_val + advance_receipts_val + contract_liabilities_val), 2)
        metrics[str(year)][1] = wc
        wc_values[year] = wc
        
        # 9. 1元收入需要的WC（元）= WC / 收入
        if revenue != "-" and revenue != 0:
            wc_per_revenue = round(wc / revenue, 2)  # 保留2位小数
            metrics[str(year)][0] = wc_per_revenue
        else:
            metrics[str(year)][0] = "-"
        
        # 10. 应收占收入比重（%）
        if accounts_receivable != "-" and revenue != "-" and revenue != 0:
            accounts_receivable_ratio = round((accounts_receivable / revenue * 100), 2)
            metrics[str(year)][9] = accounts_receivable_ratio
        else:
            metrics[str(year)][9] = "-"
        
        # 11. 预付占收入比重（%）
        if prepayment != "-" and revenue != "-" and revenue != 0:
            prepayment_ratio = round((prepayment / revenue * 100), 2)
            metrics[str(year)][10] = prepayment_ratio
        else:
            metrics[str(year)][10] = "-"
        
        # 12. 存货占收入比重（%）
        if inventory != "-" and revenue != "-" and revenue != 0:
            inventory_ratio = round((inventory / revenue * 100), 2)
            metrics[str(year)][11] = inventory_ratio
        else:
            metrics[str(year)][11] = "-"
        
        # 13. 应付占收入比重（%）
        if accounts_payable != "-" and revenue != "-" and revenue != 0:
            accounts_payable_ratio = round((accounts_payable / revenue * 100), 2)
            metrics[str(year)][12] = accounts_payable_ratio
        else:
            metrics[str(year)][12] = "-"
        
        # 14. 预收占收入比重（%）
        if advance_receipts != "-" and revenue != "-" and revenue != 0:
            advance_receipts_ratio = round((advance_receipts / revenue * 100), 2)
            metrics[str(year)][13] = advance_receipts_ratio
        else:
            metrics[str(year)][13] = "-"
        
        # 15. 合同负债占收入比重（%）
        if contract_liabilities != "-" and revenue != "-" and revenue != 0:
            contract_liabilities_ratio = round((contract_liabilities / revenue * 100), 2)
            metrics[str(year)][14] = contract_liabilities_ratio
        else:
            metrics[str(year)][14] = "-"
        
        # 16. 新增WC（亿元）= 当年WC - 上年WC
        if year == start_year:
            # 第一年没有新增WC
            metrics[str(year)][15] = "-"
        else:
            if year in wc_values and (year - 1) in wc_values:
                new_wc = round(wc_values[year] - wc_values[year - 1], 2)
                metrics[str(year)][15] = new_wc
            else:
                metrics[str(year)][15] = "-"
        
        print(f"  ✓ {year} 年数据计算完成")
    
    # 创建DataFrame
    result_df = pd.DataFrame(metrics)
    
    return result_df

def calculate_fixed_asset_metrics(symbol, start_year, end_year):
    """
    计算固定资产投入分析指标
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        包含所有固定资产投入分析指标数据的DataFrame
    """
    print("\n" + "=" * 80)
    print(f"开始计算 {symbol} 的固定资产投入分析数据（{start_year}-{end_year}）")
    print("=" * 80)
    
    # 获取数据
    data = get_annual_data(symbol, start_year, end_year)
    
    if data['balance_sheet'] is None or data['profit'] is None or data['cash_flow'] is None:
        print("✗ 资产负债表、利润表或现金流量表数据获取不完整，无法计算")
        return None
    
    # 准备结果数据
    metrics = {
        '科目': [],
    }
    
    # 初始化年份列
    for year in range(start_year, end_year + 1):
        metrics[str(year)] = []
    
    # 初始化所有科目名称
    metrics['科目'] = [
        '1元收入需要的固定资产（元）',
        '1元收入需要的长期资产（元）',
        '固定资产（亿元）',
        '长期资产（亿元）',
        '折旧（亿元）',
        '折旧/收入（%）'
    ]
    
    # 初始化所有年份的数据列表（默认值为 "-" 表示数据缺失）
    for year in range(start_year, end_year + 1):
        for _ in range(len(metrics['科目'])):
            metrics[str(year)].append("-")
    
    # 逐年计算指标
    for year in range(start_year, end_year + 1):
        print(f"\n处理 {year} 年数据...")
        
        # 获取该年份的数据
        balance_row = extract_year_data(data['balance_sheet'], year)
        profit_row = extract_year_data(data['profit'], year)
        cash_flow_row = extract_year_data(data['cash_flow'], year)
        
        if balance_row is None or profit_row is None or cash_flow_row is None:
            print(f"  ⚠ {year} 年数据缺失，该年份所有指标显示为 '-'")
            continue
        
        # 获取收入
        revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
        if revenue == "-":
            print(f"  ⚠ {year} 年营业收入数据缺失")
            continue
        
        # 计算固定资产（亿元）= 固定资产 + 在建工程 + 工程物资 - 固定资产清理
        fixed_asset = get_value_from_row(balance_row, 'FIXED_ASSET', 0)
        cip = get_value_from_row(balance_row, 'CIP', 0)
        project_material = get_value_from_row(balance_row, 'PROJECT_MATERIAL', 0)
        fixed_asset_disposal = get_value_from_row(balance_row, 'FIXED_ASSET_DISPOSAL', 0)
        
        # 处理缺失值
        if fixed_asset == "-":
            fixed_asset = 0
        if cip == "-":
            cip = 0
        if project_material == "-":
            project_material = 0
        if fixed_asset_disposal == "-":
            fixed_asset_disposal = 0
        
        total_fixed_asset = round(fixed_asset + cip + project_material - fixed_asset_disposal, 2)
        metrics[str(year)][2] = total_fixed_asset
        
        # 计算长期资产（亿元）= 固定资产 + 无形资产 + 开发支出 + 使用权资产 + 商誉 + 长期待摊费用
        intangible_asset = get_value_from_row(balance_row, 'INTANGIBLE_ASSET', 0)
        develop_expense = get_value_from_row(balance_row, 'DEVELOP_EXPENSE', 0)
        useright_asset = get_value_from_row(balance_row, 'USERIGHT_ASSET', 0)
        goodwill = get_value_from_row(balance_row, 'GOODWILL', 0)
        long_prepaid_expense = get_value_from_row(balance_row, 'LONG_PREPAID_EXPENSE', 0)
        
        # 处理缺失值
        if intangible_asset == "-":
            intangible_asset = 0
        if develop_expense == "-":
            develop_expense = 0
        if useright_asset == "-":
            useright_asset = 0
        if goodwill == "-":
            goodwill = 0
        if long_prepaid_expense == "-":
            long_prepaid_expense = 0
        
        total_long_term_asset = round(total_fixed_asset + intangible_asset + develop_expense + 
                                     useright_asset + goodwill + long_prepaid_expense, 2)
        
        # 1. 1元收入需要的固定资产（元）= 固定资产 / 收入
        if revenue != "-" and revenue != 0:
            fixed_asset_per_revenue = round(total_fixed_asset / revenue, 2)
            metrics[str(year)][0] = fixed_asset_per_revenue
        else:
            metrics[str(year)][0] = "-"
        
        # 2. 1元收入需要的长期资产（元）= 长期资产 / 收入
        if revenue != "-" and revenue != 0:
            long_term_asset_per_revenue = round(total_long_term_asset / revenue, 2)
            metrics[str(year)][1] = long_term_asset_per_revenue
        else:
            metrics[str(year)][1] = "-"
        
        # 3. 固定资产（亿元）- 已在上面计算
        # metrics[str(year)][2] = total_fixed_asset  # 已设置
        
        # 4. 长期资产（亿元）- 已在上面计算
        metrics[str(year)][3] = total_long_term_asset
        
        # 5. 折旧（亿元）- 从现金流量表获取固定资产折旧
        depreciation = get_value_from_row(cash_flow_row, 'FIXED_ASSET_DEPR', "-")
        if depreciation == "-":
            # 尝试其他可能的字段名
            depreciation = get_value_from_row(cash_flow_row, 'FA_IR_DEPR', "-")
        if depreciation == "-":
            depreciation = get_value_from_row(cash_flow_row, 'FA_DEPR', "-")
        
        metrics[str(year)][4] = depreciation
        
        # 6. 折旧/收入（%）
        if depreciation != "-" and revenue != "-" and revenue != 0:
            depreciation_ratio = round((depreciation / revenue * 100), 2)
            metrics[str(year)][5] = depreciation_ratio
        else:
            metrics[str(year)][5] = "-"
        
        print(f"  ✓ {year} 年数据计算完成")
    
    # 创建DataFrame
    result_df = pd.DataFrame(metrics)
    
    # 注意：公式说明现在通过add_formula_notes函数在Excel中单独添加，不再添加到DataFrame中
    
    return result_df

def load_employee_count_from_csv(csv_path: str) -> Dict[int, int]:
    """
    从CSV文件中加载员工数量数据
    
    参数:
        csv_path: CSV文件路径（格式：xxxx_员工数量.csv）
    
    返回:
        字典，键为年份，值为员工数量
    """
    employee_data = {}
    
    if not os.path.exists(csv_path):
        print(f"  ⚠ CSV文件不存在: {csv_path}")
        return employee_data
    
    try:
        # 尝试多种编码方式读取CSV文件
        encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'gb18030']
        df = None
        encoding_used = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(csv_path, encoding=encoding)
                encoding_used = encoding
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if df is None:
            print(f"  ⚠ 无法读取CSV文件，尝试了多种编码方式都失败: {csv_path}")
            return employee_data
        
        # 检查列名
        if '年份' not in df.columns or '员工数量' not in df.columns:
            print(f"  ⚠ CSV文件格式不正确，需要包含'年份'和'员工数量'列")
            print(f"     实际列名: {list(df.columns)}")
            return employee_data
        
        # 读取数据
        for _, row in df.iterrows():
            year = row['年份']
            count = row['员工数量']
            
            # 跳过空值
            if pd.notna(year) and pd.notna(count):
                try:
                    year_int = int(year)
                    count_int = int(count)
                    employee_data[year_int] = count_int
                except (ValueError, TypeError):
                    continue
        
        print(f"  ✓ 从CSV文件加载了 {len(employee_data)} 年的员工数量数据")
        return employee_data
        
    except Exception as e:
        print(f"  ⚠ 读取CSV文件失败: {e}")
        import traceback
        traceback.print_exc()
        return employee_data

def get_employee_count(symbol):
    """
    获取员工人数
    
    参数:
        symbol: 股票代码（不带交易所后缀）
    
    返回:
        员工人数，如果获取失败返回None
    """
    try:
        symbol_clean = symbol.replace('.SZ', '').replace('.SH', '')
        print(f"  正在获取员工人数，股票代码: {symbol_clean}")
        basic_info = ak.stock_individual_basic_info_xq(symbol=symbol_clean)
        
        if basic_info is not None and not basic_info.empty:
            print(f"  ✓ 成功获取基础信息，数据形状: {basic_info.shape}")
            print(f"  可用字段: {list(basic_info.columns)}")
            
            # 检查数据结构：如果是键值对结构（有item和value列）
            if 'item' in basic_info.columns and 'value' in basic_info.columns:
                # 在item列中查找员工人数字段
                employee_row = None
                # 优先查找 staff_num（员工人数）
                staff_num_row = basic_info[basic_info['item'] == 'staff_num']
                if not staff_num_row.empty:
                    employee_row = staff_num_row.iloc[0]
                
                if employee_row is None:
                    # 尝试查找包含"员工"和"人数"的中文字段
                    for idx, row in basic_info.iterrows():
                        item_value = str(row['item'])
                        if '员工' in item_value and '人数' in item_value:
                            employee_row = row
                            break
                
                if employee_row is None:
                    # 尝试其他可能的字段名
                    for idx, row in basic_info.iterrows():
                        item_value = str(row['item'])
                        if '人数' in item_value or 'employee' in item_value.lower() or 'staff' in item_value.lower():
                            employee_row = row
                            break
                
                if employee_row is not None:
                    value = employee_row['value']
                    print(f"  找到员工人数字段: {employee_row['item']}, 值: {value}")
                    try:
                        # 尝试转换为整数
                        if pd.notna(value) and value is not None:
                            employee_count = int(float(value))
                            print(f"  ✓ 员工人数: {employee_count}")
                            return employee_count
                        else:
                            print(f"  ⚠ 员工人数值为空")
                            return None
                    except (ValueError, TypeError):
                        print(f"  ⚠ 无法转换员工人数: {value}")
                        return None
                else:
                    print(f"  ⚠ 未找到员工人数字段")
                    # 显示所有item值，帮助调试
                    print(f"  所有item值: {basic_info['item'].tolist()}")
            else:
                # 传统列式结构（备用逻辑）
                employee_col = None
                for col in basic_info.columns:
                    if '员工' in str(col) and '人数' in str(col):
                        employee_col = col
                        break
                
                if employee_col is None:
                    for col in basic_info.columns:
                        if '人数' in str(col) or 'employee' in str(col).lower():
                            employee_col = col
                            break
                
                if employee_col is not None:
                    value = basic_info[employee_col].iloc[0]
                    print(f"  找到员工人数字段: {employee_col}, 值: {value}")
                    try:
                        employee_count = int(float(value))
                        print(f"  ✓ 员工人数: {employee_count}")
                        return employee_count
                    except (ValueError, TypeError):
                        print(f"  ⚠ 无法转换员工人数: {value}")
                        return None
                else:
                    print(f"  ⚠ 未找到员工人数字段")
                    print(f"  数据内容:\n{basic_info}")
        
        return None
    except Exception as e:
        print(f"  ⚠ 获取员工人数失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def calculate_per_capita_metrics(symbol, start_year, end_year, employee_csv_path: Optional[str] = None):
    """
    计算人均数据指标
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
        employee_csv_path: 员工数量CSV文件路径（格式：xxxx_员工数量.csv），如果提供则从CSV读取，否则使用接口
    
    返回:
        包含所有人均指标数据的DataFrame
    """
    print("\n" + "=" * 80)
    print(f"开始计算 {symbol} 的人均数据（{start_year}-{end_year}）")
    print("=" * 80)
    
    # 获取数据
    data = get_annual_data(symbol, start_year, end_year)
    
    if data['balance_sheet'] is None or data['profit'] is None or data['cash_flow'] is None:
        print("✗ 资产负债表、利润表或现金流量表数据获取不完整，无法计算")
        return None
    
    # 获取员工人数
    employee_data_by_year = {}  # 按年份存储员工数量
    
    if employee_csv_path:
        # 从CSV文件读取员工数量（按年份）
        print(f"从CSV文件读取员工数量: {employee_csv_path}")
        employee_data_by_year = load_employee_count_from_csv(employee_csv_path)
        if not employee_data_by_year:
            print("⚠ CSV文件中没有有效的员工数量数据，人均数据将显示为 '-'")
    else:
        # 使用接口获取员工人数（所有年份使用相同值）
        symbol_clean = symbol.replace('.SZ', '').replace('.SH', '')
        employee_count = get_employee_count(symbol_clean)
        if employee_count is None:
            print("⚠ 无法获取员工人数，人均数据将显示为 '-'")
        else:
            # 所有年份使用相同的员工数量
            for year in range(start_year, end_year + 1):
                employee_data_by_year[year] = employee_count
    
    # 准备结果数据
    metrics = {
        '科目': [],
    }
    
    # 初始化年份列
    for year in range(start_year, end_year + 1):
        metrics[str(year)] = []
    
    # 初始化所有科目名称
    metrics['科目'] = [
        '人数',
        '人均收入（万元）',
        '人均归母净利润（万元）',
        '人均扣非净利润（万元）',
        '人均薪酬（万元）'
    ]
    
    # 初始化所有年份的数据列表（默认值为 "-" 表示数据缺失）
    for year in range(start_year, end_year + 1):
        for _ in range(len(metrics['科目'])):
            metrics[str(year)].append("-")
    
    # 逐年计算指标
    for year in range(start_year, end_year + 1):
        print(f"\n处理 {year} 年数据...")
        
        # 获取该年份的数据
        balance_row = extract_year_data(data['balance_sheet'], year)
        profit_row = extract_year_data(data['profit'], year)
        cash_flow_row = extract_year_data(data['cash_flow'], year)
        
        if balance_row is None or profit_row is None or cash_flow_row is None:
            print(f"  ⚠ {year} 年数据缺失，该年份所有指标显示为 '-'")
            continue
        
        # 获取该年份的员工数量
        employee_count = employee_data_by_year.get(year)
        
        # 1. 人数（使用该年份的员工人数）
        if employee_count is not None:
            metrics[str(year)][0] = employee_count
        else:
            metrics[str(year)][0] = "-"
        
        # 获取收入
        revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
        if revenue == "-":
            print(f"  ⚠ {year} 年营业收入数据缺失")
            continue
        
        # 2. 人均收入（万元）= 收入（亿元） / 人数 * 10000
        if employee_count is not None and employee_count > 0 and revenue != "-":
            per_capita_revenue = round((revenue / employee_count * 10000), 2)  # 转换为万元
            metrics[str(year)][1] = per_capita_revenue
        else:
            metrics[str(year)][1] = "-"
        
        # 3. 人均归母净利润（万元）
        parent_net_profit = get_value_from_row(profit_row, 'PARENT_NETPROFIT', "-")
        if employee_count is not None and employee_count > 0 and parent_net_profit != "-":
            per_capita_profit = round((parent_net_profit / employee_count * 10000), 2)  # 转换为万元
            metrics[str(year)][2] = per_capita_profit
        else:
            metrics[str(year)][2] = "-"
        
        # 4. 人均扣非净利润（万元）
        deduct_parent_net_profit = get_value_from_row(profit_row, 'DEDUCT_PARENT_NETPROFIT', "-")
        if employee_count is not None and employee_count > 0 and deduct_parent_net_profit != "-":
            per_capita_deduct_profit = round((deduct_parent_net_profit / employee_count * 10000), 2)  # 转换为万元
            metrics[str(year)][3] = per_capita_deduct_profit
        else:
            metrics[str(year)][3] = "-"
        
        # 5. 人均薪酬（万元）
        # 获取应付职工薪酬期末余额（A）
        staff_salary_payable = get_value_from_row(balance_row, 'STAFF_SALARY_PAYABLE', 0)
        if staff_salary_payable == "-":
            staff_salary_payable = 0
        
        # 获取年初余额（B）- 上一年的期末余额
        prev_balance_row = extract_year_data(data['balance_sheet'], year - 1)
        if prev_balance_row is not None:
            prev_staff_salary_payable = get_value_from_row(prev_balance_row, 'STAFF_SALARY_PAYABLE', 0)
            if prev_staff_salary_payable == "-":
                prev_staff_salary_payable = 0
        else:
            # 如果上一年数据不存在，年初余额为0
            prev_staff_salary_payable = 0
        
        # 获取支付给职工以及为职工支付的现金（C）
        pay_staff_cash = get_value_from_row(cash_flow_row, 'PAY_STAFF_CASH', 0)
        if pay_staff_cash == "-":
            pay_staff_cash = 0
        
        # 计算员工本年度薪酬总额 = A - B + C
        total_salary = round(staff_salary_payable - prev_staff_salary_payable + pay_staff_cash, 2)
        
        # 人均薪酬 = 总薪酬 / 人数（转换为万元）
        if employee_count is not None and employee_count > 0:
            per_capita_salary = round((total_salary / employee_count * 10000), 2)  # 转换为万元
            metrics[str(year)][4] = per_capita_salary
        else:
            metrics[str(year)][4] = "-"
        
        print(f"  ✓ {year} 年数据计算完成")
    
    # 创建DataFrame
    result_df = pd.DataFrame(metrics)
    
    return result_df

def calculate_roi_metrics(symbol, start_year, end_year):
    """
    计算收益率和杜邦分析指标
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        包含所有收益率和杜邦分析指标数据的DataFrame
    """
    print("\n" + "=" * 80)
    print(f"开始计算 {symbol} 的收益率和杜邦分析数据（{start_year}-{end_year}）")
    print("=" * 80)
    
    # 获取数据
    data = get_annual_data(symbol, start_year, end_year)
    
    if data['balance_sheet'] is None or data['profit'] is None:
        print("✗ 资产负债表或利润表数据获取不完整，无法计算")
        return None
    
    # 准备结果数据
    metrics = {
        '科目': [],
    }
    
    # 初始化年份列
    for year in range(start_year, end_year + 1):
        metrics[str(year)] = []
    
    # 初始化所有科目名称
    metrics['科目'] = [
        'ROE(%)',
        'ROA(%)',
        'ROIC(%)',
        '销售净利率(%)',
        '资产周转率（次）',
        '权益乘数'
    ]
    
    # 初始化所有年份的数据列表（默认值为 "-" 表示数据缺失）
    for year in range(start_year, end_year + 1):
        for _ in range(len(metrics['科目'])):
            metrics[str(year)].append("-")
    
    # 逐年计算指标
    for year in range(start_year, end_year + 1):
        print(f"\n处理 {year} 年数据...")
        
        # 获取该年份的数据
        balance_row = extract_year_data(data['balance_sheet'], year)
        profit_row = extract_year_data(data['profit'], year)
        
        if balance_row is None or profit_row is None:
            print(f"  ⚠ {year} 年数据缺失，该年份所有指标显示为 '-'")
            continue
        
        # 获取基础数据
        # 归母净利润（亿元）
        parent_net_profit = get_value_from_row(profit_row, 'PARENT_NETPROFIT', "-")
        if parent_net_profit == "-":
            print(f"  ⚠ {year} 年归母净利润数据缺失")
            continue
        
        # 营业收入（亿元）
        revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
        if revenue == "-":
            print(f"  ⚠ {year} 年营业收入数据缺失")
            continue
        
        # 总资产（亿元）
        total_assets = get_value_from_row(balance_row, 'TOTAL_ASSETS', "-")
        if total_assets == "-":
            print(f"  ⚠ {year} 年总资产数据缺失")
            continue
        
        # 归母净资产（亿元）
        parent_equity = get_value_from_row(balance_row, 'TOTAL_PARENT_EQUITY', "-")
        if parent_equity == "-":
            print(f"  ⚠ {year} 年归母净资产数据缺失")
            continue
        
        # 1. ROE(%) = 归母净利润 / 归母净资产 * 100
        if parent_equity != "-" and parent_equity != 0:
            roe = round((parent_net_profit / parent_equity * 100), 2)
            metrics[str(year)][0] = roe
        else:
            metrics[str(year)][0] = "-"
        
        # 2. ROA(%) = 归母净利润 / 总资产 * 100
        if total_assets != "-" and total_assets != 0:
            roa = round((parent_net_profit / total_assets * 100), 2)
            metrics[str(year)][1] = roa
        else:
            metrics[str(year)][1] = "-"
        
        # 3. ROIC(%) = EBIT / 投入资本 * 100
        # EBIT = 营业利润 + 利息支出
        # 投入资本 = 总资产 - 狭义无息债务 = 股东权益 + 有息债务
        operate_profit = get_value_from_row(profit_row, 'OPERATE_PROFIT', "-")
        if operate_profit == "-":
            # 如果营业利润不存在，尝试计算：营业收入 - 营业成本 - 期间费用
            operate_cost = get_value_from_row(profit_row, 'OPERATE_COST', 0)
            if operate_cost == "-":
                operate_cost = 0
            
            # 获取期间费用
            sale_expense = get_value_from_row(profit_row, 'SALE_EXPENSE', 0)
            manage_expense = get_value_from_row(profit_row, 'MANAGE_EXPENSE', 0)
            research_expense = get_value_from_row(profit_row, 'RESEARCH_EXPENSE', 0)
            finance_expense = get_value_from_row(profit_row, 'FINANCE_EXPENSE', 0)
            
            if sale_expense == "-":
                sale_expense = 0
            if manage_expense == "-":
                manage_expense = 0
            if research_expense == "-":
                research_expense = 0
            if finance_expense == "-":
                finance_expense = 0
            
            # 计算营业利润 = 营业收入 - 营业成本 - 期间费用
            period_expenses = sale_expense + manage_expense + research_expense
            operate_profit = round(revenue - operate_cost - period_expenses, 2)
        
        # 获取利息支出（财务费用中的利息支出，如果没有则使用财务费用）
        interest_expense = get_value_from_row(profit_row, 'FE_INTEREST_EXPENSE', 0)
        if interest_expense == "-":
            # 如果财务费用中的利息支出不存在，使用财务费用（通常财务费用 = 利息支出 - 利息收入）
            finance_expense = get_value_from_row(profit_row, 'FINANCE_EXPENSE', 0)
            if finance_expense == "-":
                interest_expense = 0
            else:
                # 财务费用通常是负数（利息收入大于利息支出），取绝对值作为利息支出的近似值
                interest_expense = abs(finance_expense)
        
        # 计算EBIT = 营业利润 + 利息支出
        if operate_profit != "-":
            ebit = round(operate_profit + interest_expense, 2)
        else:
            ebit = "-"
        
        # 计算投入资本 = 总资产 - 狭义无息债务
        # 狭义无息债务 = 应付账款 + 预收账款 + 合同负债
        accounts_payable = get_value_from_row(balance_row, 'ACCOUNTS_PAYABLE', 0)
        advance_receipts = get_value_from_row(balance_row, 'ADVANCE_RECEIVABLES', 0)
        contract_liabilities = get_value_from_row(balance_row, 'CONTRACT_LIAB', 0)
        
        if accounts_payable == "-":
            accounts_payable = 0
        if advance_receipts == "-":
            advance_receipts = 0
        if contract_liabilities == "-":
            contract_liabilities = 0
        
        narrow_interest_free_debt = round(accounts_payable + advance_receipts + contract_liabilities, 2)
        
        # 投入资本 = 总资产 - 狭义无息债务
        invested_capital = round(total_assets - narrow_interest_free_debt, 2)
        
        # 计算ROIC
        if ebit != "-" and invested_capital != "-" and invested_capital != 0:
            roic = round((ebit / invested_capital * 100), 2)
            metrics[str(year)][2] = roic
        else:
            metrics[str(year)][2] = "-"
        
        # 4. 销售净利率(%) = 归母净利润 / 营业收入 * 100
        if revenue != "-" and revenue != 0:
            net_profit_margin = round((parent_net_profit / revenue * 100), 2)
            metrics[str(year)][3] = net_profit_margin
        else:
            metrics[str(year)][3] = "-"
        
        # 5. 资产周转率（次）= 营业收入 / 总资产
        if total_assets != "-" and total_assets != 0:
            asset_turnover = round(revenue / total_assets, 2)
            metrics[str(year)][4] = asset_turnover
        else:
            metrics[str(year)][4] = "-"
        
        # 6. 权益乘数 = 总资产 / 归母净资产
        if parent_equity != "-" and parent_equity != 0:
            equity_multiplier = round(total_assets / parent_equity, 2)
            metrics[str(year)][5] = equity_multiplier
        else:
            metrics[str(year)][5] = "-"
        
        print(f"  ✓ {year} 年数据计算完成")
    
    # 创建DataFrame
    result_df = pd.DataFrame(metrics)
    
    return result_df

def calculate_asset_turnover_metrics(symbol, start_year, end_year):
    """
    计算资产周转相关指标
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        包含所有资产周转指标数据的DataFrame
    """
    print("\n" + "=" * 80)
    print(f"开始计算 {symbol} 的资产周转数据（{start_year}-{end_year}）")
    print("=" * 80)
    
    # 获取数据
    data = get_annual_data(symbol, start_year, end_year)
    
    if data['balance_sheet'] is None or data['profit'] is None:
        print("✗ 资产负债表或利润表数据获取不完整，无法计算")
        return None
    
    # 准备结果数据
    metrics = {
        '科目': [],
    }
    
    # 初始化年份列
    for year in range(start_year, end_year + 1):
        metrics[str(year)] = []
    
    # 初始化所有科目名称
    metrics['科目'] = [
        '总资产（亿元）',
        '平均总资产（亿元）',
        '平均流动资产（亿元）',
        '平均存货（亿元）',
        '归母净资产（亿元）',
        '平均归母净资产（亿元）',
        '总资产周转天数',
        '流动资产周转天数',
        'WC周转天数',
        '应收周转天数',
        '存货周转天数',
        '固定资产周转天数'
    ]
    
    # 初始化所有年份的数据列表（默认值为 "-" 表示数据缺失）
    for year in range(start_year, end_year + 1):
        for _ in range(len(metrics['科目'])):
            metrics[str(year)].append("-")
    
    # 存储上一年的数据（用于计算平均值）
    prev_year_data = {}
    
    # 逐年计算指标
    for year in range(start_year, end_year + 1):
        print(f"\n处理 {year} 年数据...")
        
        # 获取该年份的数据
        balance_row = extract_year_data(data['balance_sheet'], year)
        profit_row = extract_year_data(data['profit'], year)
        
        if balance_row is None or profit_row is None:
            print(f"  ⚠ {year} 年数据缺失，该年份所有指标显示为 '-'")
            # 清空上一年的数据，因为无法计算平均值
            prev_year_data = {}
            continue
        
        # 获取营业收入
        revenue = get_value_from_row(profit_row, 'OPERATE_INCOME', "-")
        if revenue == "-":
            print(f"  ⚠ {year} 年营业收入数据缺失")
            prev_year_data = {}
            continue
        
        # 1. 总资产（亿元）
        total_assets = get_value_from_row(balance_row, 'TOTAL_ASSETS', "-")
        if total_assets == "-":
            print(f"  ⚠ {year} 年总资产数据缺失")
            prev_year_data = {}
            continue
        metrics[str(year)][0] = total_assets
        
        # 2. 平均总资产（亿元）= (上年总资产 + 当年总资产) / 2
        if (year - 1) in prev_year_data and 'total_assets' in prev_year_data[year - 1]:
            prev_total_assets = prev_year_data[year - 1]['total_assets']
            if prev_total_assets != "-":
                avg_total_assets = round((prev_total_assets + total_assets) / 2, 2)
                metrics[str(year)][1] = avg_total_assets
        else:
            # 第一年或上一年数据缺失，无法计算平均值
            metrics[str(year)][1] = "-"
        
        # 3. 平均流动资产（亿元）
        current_assets = get_value_from_row(balance_row, 'TOTAL_CURRENT_ASSETS', "-")
        if current_assets == "-":
            current_assets = 0  # 如果缺失，设为0用于计算
        
        if (year - 1) in prev_year_data and 'current_assets' in prev_year_data[year - 1]:
            prev_current_assets = prev_year_data[year - 1]['current_assets']
            if prev_current_assets != "-":
                avg_current_assets = round((prev_current_assets + current_assets) / 2, 2)
                metrics[str(year)][2] = avg_current_assets
        else:
            metrics[str(year)][2] = "-"
        
        # 4. 平均存货（亿元）
        inventory = get_value_from_row(balance_row, 'INVENTORY', "-")
        if inventory == "-":
            inventory = 0
        
        if (year - 1) in prev_year_data and 'inventory' in prev_year_data[year - 1]:
            prev_inventory = prev_year_data[year - 1]['inventory']
            if prev_inventory != "-":
                avg_inventory = round((prev_inventory + inventory) / 2, 2)
                metrics[str(year)][3] = avg_inventory
        else:
            metrics[str(year)][3] = "-"
        
        # 5. 归母净资产（亿元）
        parent_equity = get_value_from_row(balance_row, 'TOTAL_PARENT_EQUITY', "-")
        if parent_equity == "-":
            print(f"  ⚠ {year} 年归母净资产数据缺失")
            prev_year_data = {}
            continue
        metrics[str(year)][4] = parent_equity
        
        # 6. 平均归母净资产（亿元）
        if (year - 1) in prev_year_data and 'parent_equity' in prev_year_data[year - 1]:
            prev_parent_equity = prev_year_data[year - 1]['parent_equity']
            if prev_parent_equity != "-":
                avg_parent_equity = round((prev_parent_equity + parent_equity) / 2, 2)
                metrics[str(year)][5] = avg_parent_equity
        else:
            metrics[str(year)][5] = "-"
        
        # 7. 总资产周转天数 = (平均总资产 / 营业收入) × 365
        if metrics[str(year)][1] != "-" and revenue != "-" and revenue != 0:
            total_asset_turnover_days = round((metrics[str(year)][1] / revenue * 365), 2)
            metrics[str(year)][6] = total_asset_turnover_days
        else:
            metrics[str(year)][6] = "-"
        
        # 8. 流动资产周转天数 = (平均流动资产 / 营业收入) × 365
        if metrics[str(year)][2] != "-" and revenue != "-" and revenue != 0:
            current_asset_turnover_days = round((metrics[str(year)][2] / revenue * 365), 2)
            metrics[str(year)][7] = current_asset_turnover_days
        else:
            metrics[str(year)][7] = "-"
        
        # 9. WC周转天数 = (平均WC / 营业收入) × 365
        # 计算当年WC
        accounts_receivable = get_value_from_row(balance_row, 'ACCOUNTS_RECE', 0)
        prepayment = get_value_from_row(balance_row, 'PREPAYMENT', 0)
        contract_asset = get_value_from_row(balance_row, 'CONTRACT_ASSET', 0)
        accounts_payable = get_value_from_row(balance_row, 'ACCOUNTS_PAYABLE', 0)
        advance_receipts = get_value_from_row(balance_row, 'ADVANCE_RECEIVABLES', 0)
        contract_liabilities = get_value_from_row(balance_row, 'CONTRACT_LIAB', 0)
        
        if accounts_receivable == "-":
            accounts_receivable = 0
        if prepayment == "-":
            prepayment = 0
        if contract_asset == "-":
            contract_asset = 0
        if accounts_payable == "-":
            accounts_payable = 0
        if advance_receipts == "-":
            advance_receipts = 0
        if contract_liabilities == "-":
            contract_liabilities = 0
        
        wc = round((accounts_receivable + prepayment + inventory + contract_asset) - 
                   (accounts_payable + advance_receipts + contract_liabilities), 2)
        
        # 计算平均WC
        if (year - 1) in prev_year_data and 'wc' in prev_year_data[year - 1]:
            prev_wc = prev_year_data[year - 1]['wc']
            if prev_wc != "-":
                avg_wc = round((prev_wc + wc) / 2, 2)
                if revenue != "-" and revenue != 0:
                    wc_turnover_days = round((avg_wc / revenue * 365), 2)
                    metrics[str(year)][8] = wc_turnover_days
                else:
                    metrics[str(year)][8] = "-"
            else:
                metrics[str(year)][8] = "-"
        else:
            metrics[str(year)][8] = "-"
        
        # 10. 应收周转天数 = (平均应收账款 / 营业收入) × 365
        if (year - 1) in prev_year_data and 'accounts_receivable' in prev_year_data[year - 1]:
            prev_accounts_receivable = prev_year_data[year - 1]['accounts_receivable']
            if prev_accounts_receivable != "-":
                avg_accounts_receivable = round((prev_accounts_receivable + accounts_receivable) / 2, 2)
                if revenue != "-" and revenue != 0:
                    receivables_turnover_days = round((avg_accounts_receivable / revenue * 365), 2)
                    metrics[str(year)][9] = receivables_turnover_days
                else:
                    metrics[str(year)][9] = "-"
            else:
                metrics[str(year)][9] = "-"
        else:
            metrics[str(year)][9] = "-"
        
        # 11. 存货周转天数 = (平均存货 / 营业收入) × 365
        if metrics[str(year)][3] != "-" and revenue != "-" and revenue != 0:
            inventory_turnover_days = round((metrics[str(year)][3] / revenue * 365), 2)
            metrics[str(year)][10] = inventory_turnover_days
        else:
            metrics[str(year)][10] = "-"
        
        # 12. 固定资产周转天数 = (平均固定资产 / 营业收入) × 365
        # 计算当年固定资产
        fixed_asset = get_value_from_row(balance_row, 'FIXED_ASSET', 0)
        cip = get_value_from_row(balance_row, 'CIP', 0)
        project_material = get_value_from_row(balance_row, 'PROJECT_MATERIAL', 0)
        fixed_asset_disposal = get_value_from_row(balance_row, 'FIXED_ASSET_DISPOSAL', 0)
        
        if fixed_asset == "-":
            fixed_asset = 0
        if cip == "-":
            cip = 0
        if project_material == "-":
            project_material = 0
        if fixed_asset_disposal == "-":
            fixed_asset_disposal = 0
        
        total_fixed_asset = round(fixed_asset + cip + project_material - fixed_asset_disposal, 2)
        
        # 计算平均固定资产
        if (year - 1) in prev_year_data and 'fixed_asset' in prev_year_data[year - 1]:
            prev_fixed_asset = prev_year_data[year - 1]['fixed_asset']
            if prev_fixed_asset != "-":
                avg_fixed_asset = round((prev_fixed_asset + total_fixed_asset) / 2, 2)
                if revenue != "-" and revenue != 0:
                    fixed_asset_turnover_days = round((avg_fixed_asset / revenue * 365), 2)
                    metrics[str(year)][11] = fixed_asset_turnover_days
                else:
                    metrics[str(year)][11] = "-"
            else:
                metrics[str(year)][11] = "-"
        else:
            metrics[str(year)][11] = "-"
        
        # 保存当前年份的数据，供下一年使用
        prev_year_data[year] = {
            'total_assets': total_assets,
            'current_assets': current_assets if current_assets != "-" else 0,
            'inventory': inventory if inventory != "-" else 0,
            'parent_equity': parent_equity,
            'wc': wc,
            'accounts_receivable': accounts_receivable if accounts_receivable != "-" else 0,
            'fixed_asset': total_fixed_asset
        }
        
        print(f"  ✓ {year} 年数据计算完成")
    
    # 创建DataFrame
    result_df = pd.DataFrame(metrics)
    
    return result_df

def save_to_excel(df, symbol, company_name, start_year, end_year, sheet_name, output_dir="output", timestamp=None):
    """
    保存数据到Excel文件，并在数据表格下方添加公式说明区域
    
    参数:
        df: 数据框
        symbol: 股票代码
        company_name: 公司名称
        start_year: 起始年份
        end_year: 结束年份
        sheet_name: Sheet名称
        output_dir: 输出目录
        timestamp: 时间戳（格式：YYYYMMDDHHmmss），如果为None则自动生成
    """
    if df is None or df.empty:
        print(f"✗ 没有数据可保存到 {sheet_name}")
        return None
    
    # 创建输出目录
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 生成时间戳（如果未提供）
    if timestamp is None:
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    
    # 生成文件名
    symbol_clean = symbol.replace('.SZ', '').replace('.SH', '')
    filename = f"{company_name}_{start_year}-{end_year}_财务分析_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    try:
        # 如果文件已存在，追加sheet；否则创建新文件
        if os.path.exists(filepath):
            # 确保之前的文件句柄已关闭，等待一小段时间
            time.sleep(0.1)
            # 使用 openpyxl 加载工作簿，确保文件已关闭
            from openpyxl import load_workbook
            try:
                # 尝试加载工作簿，如果文件被占用会抛出异常
                wb = load_workbook(filepath)
                wb.close()
            except:
                # 如果加载失败，再等待一下
                time.sleep(0.2)
            
            # 追加模式写入
            with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✓ {sheet_name} 已追加/更新到现有Excel文件")
        else:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"✓ 已创建新Excel文件")
        
        # 确保文件句柄已关闭
        time.sleep(0.1)
        
        # 添加公式说明区域
        add_formula_notes(filepath, sheet_name, start_year, end_year)
        
        # 设置列宽自适应
        auto_adjust_column_width(filepath, sheet_name)
        
        print(f"  文件路径: {filepath}")
        print(f"  Sheet名称: {sheet_name}")
        return filepath
        
    except Exception as e:
        print(f"\n✗ 保存Excel文件失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def auto_adjust_column_width(filepath, sheet_name):
    """
    自动调整Excel sheet的列宽以适应内容
    
    参数:
        filepath: Excel文件路径
        sheet_name: Sheet名称
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        # 加载工作簿
        wb = load_workbook(filepath)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return
        
        ws = wb[sheet_name]
        
        # 遍历每一列，计算最大内容长度
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, values_only=False), start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # 遍历该列的所有单元格
            for cell in col:
                if cell.value is not None:
                    # 计算单元格内容的字符串长度
                    # 对于数字，转换为字符串；对于其他类型，直接转换为字符串
                    try:
                        cell_value = str(cell.value)
                        # 中文字符通常需要更多宽度，粗略估算：中文字符按2个字符宽度计算
                        length = 0
                        for char in cell_value:
                            if ord(char) > 127:  # 非ASCII字符（包括中文）
                                length += 2
                            else:
                                length += 1
                        if length > max_length:
                            max_length = length
                    except:
                        pass
            
            # 设置列宽（最小宽度8，最大宽度50，并添加一些padding）
            if max_length > 0:
                # 添加2个字符的padding，并限制在合理范围内
                adjusted_width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            else:
                # 如果列是空的，设置默认宽度
                ws.column_dimensions[column_letter].width = 10
        
        # 保存工作簿
        wb.save(filepath)
        wb.close()
        
    except Exception as e:
        print(f"⚠️ 自动调整列宽失败: {e}")
        import traceback
        traceback.print_exc()

def add_formula_notes(filepath, sheet_name, start_year, end_year):
    """
    在Excel sheet的数据表格下方添加公式说明区域
    
    参数:
        filepath: Excel文件路径
        sheet_name: Sheet名称
        start_year: 起始年份
        end_year: 结束年份
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        # 加载工作簿
        wb = load_workbook(filepath)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return
        
        ws = wb[sheet_name]
        
        # 找到数据表格的最后一行
        # 从最后一行向上查找，跳过可能存在的空行或公式说明行
        max_row = ws.max_row
        # 检查最后几行，如果包含"公式说明"字样，则向上查找
        for row in range(max_row, max(1, max_row - 5), -1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and '公式说明' in str(cell_value):
                max_row = row - 1
                break
        
        # 定义各sheet的公式说明
        formula_notes = {}
        
        if sheet_name == '营收基本数据':
            formula_notes = {
                '金融利润（亿元）': '金融利润 = 公允价值变动收益 + 投资收益',
                '经营利润（亿元）': '经营利润 = 归母净利润 - 金融利润',
                'CAPEX（亿元）': 'CAPEX = 购建固定资产、无形资产和其他长期资产支付的现金（来自现金流量表）'
            }
        elif sheet_name == '资产负债':
            formula_notes = {
                '狭义无息债务（亿元）': '狭义无息债务 = 应付账款 + 预收账款 + 合同负债',
                '广义无息债务（亿元）': '广义无息债务 = 应付账款 + 应付票据 + 预收账款 + 合同负债'
            }
        elif sheet_name == 'WC分析':
            formula_notes = {
                'WC（亿元）': 'WC = (应收账款 + 预付账款 + 存货 + 合同资产) - (应付账款 + 预收账款 + 合同负债)'
            }
        elif sheet_name == '固定资产投入分析':
            formula_notes = {
                '固定资产（亿元）': '固定资产 = 固定资产 + 在建工程 + 工程物资 - 固定资产清理',
                '长期资产（亿元）': '长期资产 = 固定资产 + 无形资产 + 开发支出 + 使用权资产 + 商誉 + 长期待摊费用'
            }
        elif sheet_name == '收益率和杜邦分析':
            formula_notes = {
                'ROIC(%)': 'ROIC = EBIT / 投入资本 × 100，其中EBIT = 营业利润 + 利息支出，投入资本 = 总资产 - 狭义无息债务（应付账款 + 预收账款 + 合同负债）'
            }
        
        # 如果没有公式说明，直接返回
        if not formula_notes:
            wb.close()
            return
        
        # 计算列数（科目列 + 年份列）
        num_cols = 1 + (end_year - start_year + 1)
        
        # 在数据表格下方留2行空白
        start_row = max_row + 3
        
        # 添加标题行
        title_row = start_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=num_cols)
        title_cell = ws.cell(row=title_row, column=1)
        title_cell.value = '公式说明'
        title_cell.font = Font(bold=True, size=11)
        title_cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 添加公式说明行
        current_row = start_row + 1
        for metric_name, formula in formula_notes.items():
            # 合并第一列（科目列）和所有年份列
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
            formula_cell = ws.cell(row=current_row, column=1)
            formula_cell.value = f'{metric_name}: {formula}'
            formula_cell.font = Font(size=10)
            formula_cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            formula_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            current_row += 1
        
        # 保存工作簿
        wb.save(filepath)
        wb.close()
        
    except Exception as e:
        print(f"⚠️ 添加公式说明失败: {e}")
        import traceback
        traceback.print_exc()

def main():
    # 指定股票代码和年份范围
    symbol = "603486"  # 可以修改为其他股票代码
    start_year = 2015  # 起始年份
    end_year = 2024     # 结束年份
    
    # 员工数量CSV文件路径（可选，如果提供则从CSV读取，否则使用接口）
    # 格式：xxxx_员工数量.csv，例如：600728_员工数量.csv
    employee_csv_path = r"E:\stock\行业\科沃斯\603486_员工数量.csv"  # 例如：r"G:\移动云盘同步文件夹\13600004997\生活\投资\资料\财报\佳都\600728_员工数量.csv"
    
    # 生成时间戳（用于文件名，确保所有sheet保存到同一个文件）
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    
    # 获取股票名称
    company_name = get_symbol_name(symbol)
    print(f"股票代码: {symbol}")
    print(f"公司名称: {company_name}")
    print(f"分析年份: {start_year}-{end_year}\n")
    
    # 计算营收基本数据
    result_df = calculate_revenue_metrics(symbol, start_year, end_year)
    
    if result_df is not None and not result_df.empty:
        # 打印到控制台
        print("\n" + "=" * 80)
        print(f"营收基本数据（{start_year}-{end_year}）")
        print("=" * 80)
        print(result_df.to_string(index=False))
        
        # 保存到Excel
        save_to_excel(result_df, symbol, company_name, start_year, end_year, '营收基本数据', timestamp=timestamp)
    else:
        print("\n✗ 未能生成营收基本数据")
    
    # 计算费用构成数据
    expense_df = calculate_expense_metrics(symbol, start_year, end_year)
    
    if expense_df is not None and not expense_df.empty:
        # 打印到控制台
        print("\n" + "=" * 80)
        print(f"费用构成数据（{start_year}-{end_year}）")
        print("=" * 80)
        print(expense_df.to_string(index=False))
        
        # 保存到Excel（追加到同一个文件）
        save_to_excel(expense_df, symbol, company_name, start_year, end_year, '费用构成', timestamp=timestamp)
    else:
        print("\n✗ 未能生成费用构成数据")
    
    # 计算增长率数据
    growth_df = calculate_growth_metrics(symbol, start_year, end_year)
    
    if growth_df is not None and not growth_df.empty:
        # 打印到控制台
        print("\n" + "=" * 80)
        print(f"增长率数据（{start_year}-{end_year}）")
        print("=" * 80)
        print(growth_df.to_string(index=False))
        
        # 保存到Excel（追加到同一个文件）
        save_to_excel(growth_df, symbol, company_name, start_year, end_year, '增长', timestamp=timestamp)
    else:
        print("\n✗ 未能生成增长率数据")
    
    # 计算资产负债数据
    balance_df = calculate_balance_sheet_metrics(symbol, start_year, end_year)
    
    if balance_df is not None and not balance_df.empty:
        # 打印到控制台
        print("\n" + "=" * 80)
        print(f"资产负债数据（{start_year}-{end_year}）")
        print("=" * 80)
        print(balance_df.to_string(index=False))
        
        # 保存到Excel（追加到同一个文件）
        save_to_excel(balance_df, symbol, company_name, start_year, end_year, '资产负债', timestamp=timestamp)
    else:
        print("\n✗ 未能生成资产负债数据")
    
    # 计算WC分析数据
    wc_df = calculate_wc_metrics(symbol, start_year, end_year)
    
    if wc_df is not None and not wc_df.empty:
        # 打印到控制台
        print("\n" + "=" * 80)
        print(f"WC分析数据（{start_year}-{end_year}）")
        print("=" * 80)
        print(wc_df.to_string(index=False))
        
        # 保存到Excel（追加到同一个文件）
        save_to_excel(wc_df, symbol, company_name, start_year, end_year, 'WC分析', timestamp=timestamp)
    else:
        print("\n✗ 未能生成WC分析数据")
    
    # 计算固定资产投入分析数据
    fixed_asset_df = calculate_fixed_asset_metrics(symbol, start_year, end_year)
    
    if fixed_asset_df is not None and not fixed_asset_df.empty:
        # 打印到控制台
        print("\n" + "=" * 80)
        print(f"固定资产投入分析数据（{start_year}-{end_year}）")
        print("=" * 80)
        print(fixed_asset_df.to_string(index=False))
        
        # 保存到Excel（追加到同一个文件）
        save_to_excel(fixed_asset_df, symbol, company_name, start_year, end_year, '固定资产投入分析', timestamp=timestamp)
    else:
        print("\n✗ 未能生成固定资产投入分析数据")
    
    # 计算收益率和杜邦分析数据
    roi_df = calculate_roi_metrics(symbol, start_year, end_year)
    
    if roi_df is not None and not roi_df.empty:
        # 打印到控制台
        print("\n" + "=" * 80)
        print(f"收益率和杜邦分析数据（{start_year}-{end_year}）")
        print("=" * 80)
        print(roi_df.to_string(index=False))
        
        # 保存到Excel（追加到同一个文件）
        save_to_excel(roi_df, symbol, company_name, start_year, end_year, '收益率和杜邦分析', timestamp=timestamp)
    else:
        print("\n✗ 未能生成收益率和杜邦分析数据")
    
    # 计算资产周转数据
    asset_turnover_df = calculate_asset_turnover_metrics(symbol, start_year, end_year)
    
    if asset_turnover_df is not None and not asset_turnover_df.empty:
        # 打印到控制台
        print("\n" + "=" * 80)
        print(f"资产周转数据（{start_year}-{end_year}）")
        print("=" * 80)
        print(asset_turnover_df.to_string(index=False))
        
        # 保存到Excel（追加到同一个文件）
        save_to_excel(asset_turnover_df, symbol, company_name, start_year, end_year, '资产周转', timestamp=timestamp)
    else:
        print("\n✗ 未能生成资产周转数据")
    
    # 计算人均数据
    per_capita_df = calculate_per_capita_metrics(symbol, start_year, end_year, employee_csv_path=employee_csv_path)
    
    if per_capita_df is not None and not per_capita_df.empty:
        # 打印到控制台
        print("\n" + "=" * 80)
        print(f"人均数据（{start_year}-{end_year}）")
        print("=" * 80)
        print(per_capita_df.to_string(index=False))
        
        # 保存到Excel（追加到同一个文件）
        save_to_excel(per_capita_df, symbol, company_name, start_year, end_year, '人均数据', timestamp=timestamp)
    else:
        print("\n✗ 未能生成人均数据")
    
    # 最终提示
    symbol_clean = symbol.replace('.SZ', '').replace('.SH', '')
    filename = f"{company_name}_{start_year}-{end_year}_财务分析_{timestamp}.xlsx"
    filepath = os.path.join("output", filename)
    if os.path.exists(filepath):
        print(f"\n{'='*80}")
        print(f"✓ 所有数据已保存到: {filepath}")
        print(f"{'='*80}")

if __name__ == "__main__":
    main()

