# -*- coding: utf-8 -*-
"""
ç»Ÿä¸€è´¢åŠ¡å·¥å…·ï¼ˆAè‚¡ + æ¸¯è‚¡ï¼‰

æ•´åˆä¸‰å¤§åŠŸèƒ½ï¼š
1) è´¢åŠ¡åˆ†æï¼ˆ9ä¸ªSheetï¼ŒAè‚¡/æ¸¯è‚¡ï¼‰
2) å†å¹´ä¸‰å¤§æŠ¥è¡¨ä¸‹è½½ï¼ˆèµ„äº§è´Ÿå€ºè¡¨ã€åˆ©æ¶¦è¡¨ã€ç°é‡‘æµé‡è¡¨ï¼‰
3) å†å¹´å‘˜å·¥æ•°é‡æå–ï¼ˆå¹´æŠ¥PDFï¼‰

è¯´æ˜ï¼š
- Aè‚¡é»˜è®¤è´§å¸ï¼šäººæ°‘å¸ï¼ˆCNYï¼‰
- æ¸¯è‚¡é»˜è®¤è´§å¸ï¼šæ¸¯å¸ï¼ˆHKDï¼‰ï¼Œä¸è¿›è¡Œè´§å¸è½¬æ¢
"""

import os
import io
import sys
import importlib.util
from datetime import datetime
from typing import Dict, Optional

import pandas as pd
import streamlit as st
import plotly.express as px

# -----------------------------
# å·¥å…·ï¼šåŠ¨æ€å¯¼å…¥æ¨¡å—
# -----------------------------
def load_module(name: str, path: str):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module

# é¢„åŠ è½½æ ¸å¿ƒæ¨¡å—
fa_a = load_module("financial_analysis_a", "07_è´¢åŠ¡åˆ†æ.py")
fa_hk = load_module("financial_analysis_hk", "hk_financial_analysis_full.py")
hk_adapter = load_module("hk_adapter", "hk_financial_adapter.py")
dl_tool = load_module("report_downloader", "è´¢åŠ¡æŠ¥è¡¨ä¸‹è½½å·¥å…·.py")
emp_a = load_module("emp_a", "æ™ºèƒ½_ä»å¹´æŠ¥æå–å‘˜å·¥æ•°é‡.py")
emp_hk = load_module("emp_hk", "æ¸¯è‚¡_ä»å¹´æŠ¥æå–å‘˜å·¥æ•°é‡.py")

# -----------------------------
# é¡µé¢é…ç½®
# -----------------------------
st.set_page_config(
    page_title="ç»Ÿä¸€è´¢åŠ¡å·¥å…·ï¼ˆAè‚¡+æ¸¯è‚¡ï¼‰",
    page_icon="ğŸ“Š",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -----------------------------
# ä¾§è¾¹æ ï¼šå¸‚åœºä¸åŠŸèƒ½é€‰æ‹©
# -----------------------------
st.sidebar.header("ğŸŒ å¸‚åœºä¸åŠŸèƒ½")
market = st.sidebar.radio("é€‰æ‹©å¸‚åœº", ["Aè‚¡", "æ¸¯è‚¡"], horizontal=True)

# åŠŸèƒ½åˆ—è¡¨ï¼ˆAè‚¡å’Œæ¸¯è‚¡éƒ½æ”¯æŒå¹´æŠ¥PDFä¸‹è½½ï¼‰
feature_options = ["ğŸ“Š è´¢åŠ¡åˆ†æ", "ğŸ“„ æŠ¥è¡¨ä¸‹è½½", "ğŸ‘¥ å‘˜å·¥æ•°é‡æå–", "ğŸ“¥ å¹´æŠ¥PDFä¸‹è½½"]

feature = st.sidebar.radio("é€‰æ‹©åŠŸèƒ½", feature_options)

# å…¬å…±è¾“å…¥ï¼šè‚¡ç¥¨ä»£ç ä¸å¹´ä»½
st.sidebar.markdown("---")
symbol_help = "Aè‚¡ï¼š6ä½ä»£ç ï¼Œå¦‚ 603486ï¼›æ¸¯è‚¡ï¼š5ä½ä»£ç ï¼Œå¦‚ 00700"
symbol_default = "603486" if market == "Aè‚¡" else "00700"
symbol = st.sidebar.text_input("è‚¡ç¥¨ä»£ç ", value=symbol_default, help=symbol_help)

col_y1, col_y2 = st.sidebar.columns(2)
with col_y1:
    start_year = st.number_input("èµ·å§‹å¹´ä»½", min_value=2000, max_value=2035, value=2020, step=1)
with col_y2:
    end_year = st.number_input("ç»“æŸå¹´ä»½", min_value=2000, max_value=2035, value=2024, step=1)
if start_year > end_year:
    st.sidebar.error("èµ·å§‹å¹´ä»½ä¸èƒ½å¤§äºç»“æŸå¹´ä»½")
    st.stop()

# -----------------------------
# è¾…åŠ©å‡½æ•°ï¼šè·å–å…¬å¼è¯´æ˜
# -----------------------------
def get_formula_notes(sheet_name):
    """
    è·å–æŒ‡å®šsheetçš„å…¬å¼è¯´æ˜
    
    å‚æ•°:
        sheet_name: Sheetåç§°
    
    è¿”å›:
        å­—å…¸ï¼Œæ ¼å¼ä¸º {æŒ‡æ ‡åç§°: å…¬å¼è¯´æ˜}
    """
    formula_notes = {}
    
    if sheet_name == 'è¥æ”¶åŸºæœ¬æ•°æ®':
        formula_notes = {
            'é‡‘èåˆ©æ¶¦ï¼ˆäº¿å…ƒï¼‰': 'é‡‘èåˆ©æ¶¦ = å…¬å…ä»·å€¼å˜åŠ¨æ”¶ç›Š + æŠ•èµ„æ”¶ç›Š',
            'ç»è¥åˆ©æ¶¦ï¼ˆäº¿å…ƒï¼‰': 'ç»è¥åˆ©æ¶¦ = å½’æ¯å‡€åˆ©æ¶¦ - é‡‘èåˆ©æ¶¦',
            'CAPEXï¼ˆäº¿å…ƒï¼‰': 'CAPEX = è´­å»ºå›ºå®šèµ„äº§ã€æ— å½¢èµ„äº§å’Œå…¶ä»–é•¿æœŸèµ„äº§æ”¯ä»˜çš„ç°é‡‘ï¼ˆæ¥è‡ªç°é‡‘æµé‡è¡¨ï¼‰'
        }
    elif sheet_name == 'èµ„äº§è´Ÿå€º':
        formula_notes = {
            'ç‹­ä¹‰æ— æ¯å€ºåŠ¡ï¼ˆäº¿å…ƒï¼‰': 'ç‹­ä¹‰æ— æ¯å€ºåŠ¡ = åº”ä»˜è´¦æ¬¾ + é¢„æ”¶è´¦æ¬¾ + åˆåŒè´Ÿå€º',
            'å¹¿ä¹‰æ— æ¯å€ºåŠ¡ï¼ˆäº¿å…ƒï¼‰': 'å¹¿ä¹‰æ— æ¯å€ºåŠ¡ = åº”ä»˜è´¦æ¬¾ + åº”ä»˜ç¥¨æ® + é¢„æ”¶è´¦æ¬¾ + åˆåŒè´Ÿå€º'
        }
    elif sheet_name == 'WCåˆ†æ':
        formula_notes = {
            'WCï¼ˆäº¿å…ƒï¼‰': 'WC = (åº”æ”¶è´¦æ¬¾ + é¢„ä»˜è´¦æ¬¾ + å­˜è´§ + åˆåŒèµ„äº§) - (åº”ä»˜è´¦æ¬¾ + é¢„æ”¶è´¦æ¬¾ + åˆåŒè´Ÿå€º)'
        }
    elif sheet_name == 'å›ºå®šèµ„äº§æŠ•å…¥åˆ†æ':
        formula_notes = {
            'å›ºå®šèµ„äº§ï¼ˆäº¿å…ƒï¼‰': 'å›ºå®šèµ„äº§ = å›ºå®šèµ„äº§ + åœ¨å»ºå·¥ç¨‹ + å·¥ç¨‹ç‰©èµ„ - å›ºå®šèµ„äº§æ¸…ç†',
            'é•¿æœŸèµ„äº§ï¼ˆäº¿å…ƒï¼‰': 'é•¿æœŸèµ„äº§ = å›ºå®šèµ„äº§ + æ— å½¢èµ„äº§ + å¼€å‘æ”¯å‡º + ä½¿ç”¨æƒèµ„äº§ + å•†èª‰ + é•¿æœŸå¾…æ‘Šè´¹ç”¨'
        }
    elif sheet_name == 'æ”¶ç›Šç‡å’Œæœé‚¦åˆ†æ':
        formula_notes = {
            'ROIC(%)': 'ROIC = EBIT / æŠ•å…¥èµ„æœ¬ Ã— 100ï¼Œå…¶ä¸­EBIT = è¥ä¸šåˆ©æ¶¦ + åˆ©æ¯æ”¯å‡ºï¼ŒæŠ•å…¥èµ„æœ¬ = æ€»èµ„äº§ - ç‹­ä¹‰æ— æ¯å€ºåŠ¡ï¼ˆåº”ä»˜è´¦æ¬¾ + é¢„æ”¶è´¦æ¬¾ + åˆåŒè´Ÿå€ºï¼‰'
        }
    
    return formula_notes

# -----------------------------
# åŠŸèƒ½ 1ï¼šè´¢åŠ¡åˆ†æ
# -----------------------------
def run_financial_analysis():
    st.header("ğŸ“Š è´¢åŠ¡åˆ†æç»“æœ")
    # åˆ†ææ¨¡å—é€‰æ‹©
    st.sidebar.markdown("### ğŸ“‹ åˆ†ææ¨¡å—")
    modules = {
        "è¥æ”¶åŸºæœ¬æ•°æ®": st.sidebar.checkbox("è¥æ”¶åŸºæœ¬æ•°æ®", value=True),
        "è´¹ç”¨æ„æˆ": st.sidebar.checkbox("è´¹ç”¨æ„æˆ", value=True),
        "å¢é•¿ç‡": st.sidebar.checkbox("å¢é•¿ç‡", value=True),
        "èµ„äº§è´Ÿå€º": st.sidebar.checkbox("èµ„äº§è´Ÿå€º", value=True),
        "WCåˆ†æ": st.sidebar.checkbox("WCåˆ†æ", value=True),
        "å›ºå®šèµ„äº§æŠ•å…¥åˆ†æ": st.sidebar.checkbox("å›ºå®šèµ„äº§æŠ•å…¥åˆ†æ", value=True),
        "æ”¶ç›Šç‡å’Œæœé‚¦åˆ†æ": st.sidebar.checkbox("æ”¶ç›Šç‡å’Œæœé‚¦åˆ†æ", value=True),
        "èµ„äº§å‘¨è½¬": st.sidebar.checkbox("èµ„äº§å‘¨è½¬", value=True),
        "äººå‡æ•°æ®": st.sidebar.checkbox("äººå‡æ•°æ®", value=True),
    }
    # å¯é€‰å‘˜å·¥CSV
    employee_csv = st.sidebar.file_uploader("å‘˜å·¥æ•°é‡CSVï¼ˆå¯é€‰ï¼Œå¹´ä»½,å‘˜å·¥æ•°é‡ï¼‰", type=["csv"])

    run_btn = st.sidebar.button("ğŸš€ å¼€å§‹åˆ†æ", type="primary", use_container_width=True)
    
    # å…ˆæ£€æŸ¥æ˜¯å¦æœ‰å·²ä¿å­˜çš„ç»“æœ
    session_key = f"analysis_results_{market}_{symbol}_{start_year}_{end_year}"
    
    # å¦‚æœç‚¹å‡»äº†æŒ‰é’®ï¼Œæ‰§è¡Œåˆ†æ
    if run_btn:
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        results: Dict[str, pd.DataFrame] = {}

        try:
            if market == "Aè‚¡":
                fa = fa_a
            else:
                fa = fa_hk
            company_name = fa.get_symbol_name(symbol) if hasattr(fa, "get_symbol_name") else symbol

            progress = st.progress(0)
            done = 0
            total = sum(modules.values()) or 1

            def step():
                nonlocal done
                done += 1
                progress.progress(min(1.0, done / total))

            # è¥æ”¶åŸºæœ¬æ•°æ®
            if modules["è¥æ”¶åŸºæœ¬æ•°æ®"]:
                df = fa.calculate_revenue_metrics(symbol, start_year, end_year)
                if df is not None and not df.empty:
                    results["è¥æ”¶åŸºæœ¬æ•°æ®"] = df
                    fa.save_to_excel(df, symbol, company_name, start_year, end_year, "è¥æ”¶åŸºæœ¬æ•°æ®", timestamp=timestamp)
                step()
            # è´¹ç”¨æ„æˆ
            if modules["è´¹ç”¨æ„æˆ"]:
                df = fa.calculate_expense_metrics(symbol, start_year, end_year)
                if df is not None and not df.empty:
                    results["è´¹ç”¨æ„æˆ"] = df
                    fa.save_to_excel(df, symbol, company_name, start_year, end_year, "è´¹ç”¨æ„æˆ", timestamp=timestamp)
                step()
            # å¢é•¿ç‡
            if modules["å¢é•¿ç‡"]:
                df = fa.calculate_growth_metrics(symbol, start_year, end_year)
                if df is not None and not df.empty:
                    results["å¢é•¿"] = df
                    fa.save_to_excel(df, symbol, company_name, start_year, end_year, "å¢é•¿", timestamp=timestamp)
                step()
            # èµ„äº§è´Ÿå€º
            if modules["èµ„äº§è´Ÿå€º"]:
                df = fa.calculate_balance_sheet_metrics(symbol, start_year, end_year)
                if df is not None and not df.empty:
                    results["èµ„äº§è´Ÿå€º"] = df
                    fa.save_to_excel(df, symbol, company_name, start_year, end_year, "èµ„äº§è´Ÿå€º", timestamp=timestamp)
                step()
            # WCåˆ†æ
            if modules["WCåˆ†æ"]:
                df = fa.calculate_wc_metrics(symbol, start_year, end_year)
                if df is not None and not df.empty:
                    results["WCåˆ†æ"] = df
                    fa.save_to_excel(df, symbol, company_name, start_year, end_year, "WCåˆ†æ", timestamp=timestamp)
                step()
            # å›ºå®šèµ„äº§æŠ•å…¥åˆ†æ
            if modules["å›ºå®šèµ„äº§æŠ•å…¥åˆ†æ"]:
                df = fa.calculate_fixed_asset_metrics(symbol, start_year, end_year)
                if df is not None and not df.empty:
                    results["å›ºå®šèµ„äº§æŠ•å…¥åˆ†æ"] = df
                    fa.save_to_excel(df, symbol, company_name, start_year, end_year, "å›ºå®šèµ„äº§æŠ•å…¥åˆ†æ", timestamp=timestamp)
                step()
            # æ”¶ç›Šç‡å’Œæœé‚¦åˆ†æ
            if modules["æ”¶ç›Šç‡å’Œæœé‚¦åˆ†æ"]:
                df = fa.calculate_roi_metrics(symbol, start_year, end_year)
                if df is not None and not df.empty:
                    results["æ”¶ç›Šç‡å’Œæœé‚¦åˆ†æ"] = df
                    fa.save_to_excel(df, symbol, company_name, start_year, end_year, "æ”¶ç›Šç‡å’Œæœé‚¦åˆ†æ", timestamp=timestamp)
                step()
            # èµ„äº§å‘¨è½¬
            if modules["èµ„äº§å‘¨è½¬"]:
                df = fa.calculate_asset_turnover_metrics(symbol, start_year, end_year)
                if df is not None and not df.empty:
                    results["èµ„äº§å‘¨è½¬"] = df
                    fa.save_to_excel(df, symbol, company_name, start_year, end_year, "èµ„äº§å‘¨è½¬", timestamp=timestamp)
                step()
            # äººå‡æ•°æ®
            if modules["äººå‡æ•°æ®"]:
                csv_path = None
                if employee_csv:
                    import tempfile
                    csv_bytes = employee_csv.getvalue()
                    # ä½¿ç”¨ç³»ç»Ÿä¸´æ—¶ç›®å½•ï¼Œå…¼å®¹ Windows å’Œ Linux
                    temp_dir = tempfile.gettempdir()
                    csv_path = os.path.join(temp_dir, employee_csv.name)
                    with open(csv_path, "wb") as f:
                        f.write(csv_bytes)
                df = fa.calculate_per_capita_metrics(symbol, start_year, end_year, employee_csv_path=csv_path)
                if df is not None and not df.empty:
                    results["äººå‡æ•°æ®"] = df
                    fa.save_to_excel(df, symbol, company_name, start_year, end_year, "äººå‡æ•°æ®", timestamp=timestamp)
                step()

            progress.progress(1.0)
            st.success("åˆ†æå®Œæˆï¼")

            if not results:
                st.warning("æœªç”Ÿæˆä»»ä½•ç»“æœï¼Œè¯·æ£€æŸ¥æ•°æ®æ˜¯å¦å¯ç”¨ã€‚")
                return

            # ä¿å­˜ç»“æœåˆ° session_state
            st.session_state[session_key] = results
            st.session_state[f"{session_key}_company"] = company_name
            st.session_state[f"{session_key}_timestamp"] = timestamp
            # æ³¨æ„ï¼šsave_to_excelå‡½æ•°ç»Ÿä¸€ä½¿ç”¨"è´¢åŠ¡åˆ†æ"ä½œä¸ºæ–‡ä»¶åï¼Œä¸åŒºåˆ†Aè‚¡/æ¸¯è‚¡
            st.session_state[f"{session_key}_filepath"] = os.path.join("output", f"{company_name}_{start_year}-{end_year}_è´¢åŠ¡åˆ†æ_{timestamp}.xlsx")

        except Exception as e:
            st.error(f"åˆ†æå¤±è´¥ï¼š{e}")
            import traceback
            st.code(traceback.format_exc())

    # æ˜¾ç¤ºåˆ†æç»“æœï¼ˆä» session_state è¯»å–ï¼‰
    if session_key in st.session_state:
        results = st.session_state[session_key]
        company_name = st.session_state.get(f"{session_key}_company", symbol)
        timestamp = st.session_state.get(f"{session_key}_timestamp", "")
        filepath = st.session_state.get(f"{session_key}_filepath", "")

        sheet = st.selectbox("é€‰æ‹©è¦æŸ¥çœ‹çš„Sheet", list(results.keys()))
        
        # æ˜¾ç¤ºæ•°æ®è¡¨æ ¼
        st.subheader(f"ğŸ“Š {sheet}")
        # å°†DataFrameè½¬æ¢ä¸ºå­—ç¬¦ä¸²ç±»å‹ä»¥é¿å…PyArrowç±»å‹è½¬æ¢é—®é¢˜ï¼ˆæ··åˆç±»å‹ï¼šæ•°å€¼å’Œ"-"ï¼‰
        display_df = results[sheet].astype(str)
        st.dataframe(display_df, width='stretch', height=420)
        
        # æ˜¾ç¤ºå…¬å¼æ³¨é‡Š
        formula_notes = get_formula_notes(sheet)
        if formula_notes:
            st.markdown("---")
            st.subheader("ğŸ“ å…¬å¼è¯´æ˜")
            for metric_name, formula in formula_notes.items():
                st.markdown(f"**{metric_name}**: {formula}")
        
        # è¶‹åŠ¿å›¾ï¼ˆä»…æ˜¾ç¤ºæ•°å€¼æŒ‡æ ‡ï¼‰
        try:
            df = results[sheet]
            # DataFrame æ ¼å¼ï¼šç¬¬ä¸€åˆ—æ˜¯"ç§‘ç›®"ï¼Œå…¶ä»–åˆ—æ˜¯å¹´ä»½ï¼ˆå¦‚'2020', '2021'ç­‰ï¼‰
            if "ç§‘ç›®" in df.columns:
                # è·å–æ‰€æœ‰å¹´ä»½åˆ—ï¼ˆæ•°å­—å­—ç¬¦ä¸²ï¼‰
                year_cols = [col for col in df.columns if col != "ç§‘ç›®" and col.isdigit()]
                
                if len(year_cols) >= 2:  # è‡³å°‘éœ€è¦2å¹´æ•°æ®æ‰èƒ½ç”»è¶‹åŠ¿
                    st.subheader("ğŸ“ˆ è¶‹åŠ¿å›¾")
                    
                    # è·å–æ‰€æœ‰ç§‘ç›®ï¼ˆæŒ‡æ ‡ï¼‰
                    all_metrics = df["ç§‘ç›®"].tolist()
                    
                    # ç”¨æˆ·é€‰æ‹©è¦æ˜¾ç¤ºçš„æŒ‡æ ‡
                    selected_metrics = st.multiselect(
                        "é€‰æ‹©è¦å¯è§†åŒ–çš„æŒ‡æ ‡",
                        options=all_metrics,
                        default=all_metrics[:min(3, len(all_metrics))],  # é»˜è®¤é€‰å‰3ä¸ª
                        key=f"chart_metrics_{sheet}"
                    )
                    
                    if selected_metrics:
                        # å‡†å¤‡ç»˜å›¾æ•°æ®
                        chart_data = []
                        for metric in selected_metrics:
                            metric_row = df[df["ç§‘ç›®"] == metric]
                            if not metric_row.empty:
                                for year_col in year_cols:
                                    value = metric_row[year_col].iloc[0]
                                    # è·³è¿‡ "-" å’Œéæ•°å€¼
                                    if value != "-" and pd.notna(value):
                                        try:
                                            numeric_value = float(value)
                                            chart_data.append({
                                                "å¹´ä»½": int(year_col),
                                                "æŒ‡æ ‡": metric,
                                                "æ•°å€¼": numeric_value
                                            })
                                        except (ValueError, TypeError):
                                            pass
                        
                        if chart_data:
                            chart_df = pd.DataFrame(chart_data)
                            fig = px.line(
                                chart_df,
                                x="å¹´ä»½",
                                y="æ•°å€¼",
                                color="æŒ‡æ ‡",
                                markers=True,
                                title=f"{sheet} - è¶‹åŠ¿å›¾"
                            )
                            fig.update_layout(hovermode="x unified")
                            st.plotly_chart(fig, use_container_width=True)
                        else:
                            st.info("ğŸ’¡ æ‰€é€‰æŒ‡æ ‡æ²¡æœ‰å¯ç»˜åˆ¶çš„æ•°å€¼æ•°æ®")
        except Exception as e:
            st.warning(f"âš ï¸ å›¾è¡¨ç”Ÿæˆå¤±è´¥ï¼š{str(e)}")
            import traceback
            st.code(traceback.format_exc())

        # ä¸‹è½½Excelæ–‡ä»¶
        if filepath and os.path.exists(filepath):
            with open(filepath, "rb") as f:
                st.download_button("ğŸ“¥ ä¸‹è½½Excelæ–‡ä»¶", data=f.read(), file_name=os.path.basename(filepath),
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Excel æ–‡ä»¶å°šæœªç”Ÿæˆæˆ–è·¯å¾„ä¸å­˜åœ¨ã€‚")
    else:
        # æ²¡æœ‰ç»“æœï¼Œæç¤ºç”¨æˆ·
        st.info("åœ¨å·¦ä¾§é€‰æ‹©åˆ†ææ¨¡å—å¹¶ç‚¹å‡»ã€å¼€å§‹åˆ†æã€‘æŒ‰é’®ã€‚")

# -----------------------------
# åŠŸèƒ½ 2ï¼šæŠ¥è¡¨ä¸‹è½½
# -----------------------------
def run_report_download():
    st.header("ğŸ“„ è´¢åŠ¡æŠ¥è¡¨ä¸‹è½½")
    st.sidebar.markdown("### ğŸ“‹ æŠ¥è¡¨ç±»å‹")
    dl_balance = st.sidebar.checkbox("èµ„äº§è´Ÿå€ºè¡¨", value=True)
    dl_profit = st.sidebar.checkbox("åˆ©æ¶¦è¡¨", value=True)
    dl_cash = st.sidebar.checkbox("ç°é‡‘æµé‡è¡¨", value=True)
    run_btn = st.sidebar.button("ğŸš€ å¼€å§‹ä¸‹è½½", type="primary", use_container_width=True)
    
    # åˆå§‹åŒ– session_state
    session_key = f"report_data_{market}_{symbol}_{start_year}_{end_year}"
    if run_btn:
        # ç‚¹å‡»æŒ‰é’®æ—¶æ¸…ç©ºæ—§æ•°æ®
        if session_key in st.session_state:
            del st.session_state[session_key]
        if f"{session_key}_excel" in st.session_state:
            del st.session_state[f"{session_key}_excel"]
        if f"{session_key}_company" in st.session_state:
            del st.session_state[f"{session_key}_company"]
    
    # å¦‚æœå·²æœ‰æ•°æ®ï¼Œç›´æ¥ä½¿ç”¨ï¼›å¦åˆ™æç¤ºç”¨æˆ·ç‚¹å‡»æŒ‰é’®
    if session_key not in st.session_state:
        if not run_btn:
            st.info("åœ¨å·¦ä¾§é€‰æ‹©æŠ¥è¡¨ç±»å‹å¹¶ç‚¹å‡»å¼€å§‹ä¸‹è½½ã€‚")
            return
    else:
        # å·²æœ‰æ•°æ®ï¼Œæ˜¾ç¤ºé¢„è§ˆå’Œä¸‹è½½
        result_dict = st.session_state[session_key]
        excel_bytes = st.session_state.get(f"{session_key}_excel")
        company_name = st.session_state.get(f"{session_key}_company", symbol)
        
        # æ˜¾ç¤ºé¢„è§ˆé€‰æ‹©å™¨ï¼ˆAè‚¡å’Œæ¸¯è‚¡éƒ½æ”¯æŒï¼‰
        st.subheader(f"ğŸ“Š {'Aè‚¡' if market == 'Aè‚¡' else 'æ¸¯è‚¡'}æŠ¥è¡¨é¢„è§ˆ")
        year_options = sorted(result_dict.keys())
        if year_options:
            col_y, col_t = st.columns(2)
            with col_y:
                sel_year = st.selectbox("é€‰æ‹©å¹´ä»½", year_options, key=f"report_year_{market}")
            with col_t:
                sel_type = st.selectbox("é€‰æ‹©æŠ¥è¡¨ç±»å‹", ["èµ„äº§è´Ÿå€ºè¡¨", "åˆ©æ¶¦è¡¨", "ç°é‡‘æµé‡è¡¨"], key=f"report_type_{market}")
            
            stmt_map = {
                "èµ„äº§è´Ÿå€ºè¡¨": "balance",
                "åˆ©æ¶¦è¡¨": "profit",
                "ç°é‡‘æµé‡è¡¨": "cash_flow",
            }
            df_preview = result_dict.get(sel_year, {}).get(stmt_map[sel_type])
            if df_preview is None or df_preview.empty:
                st.info(f"{sel_year} å¹´çš„ {sel_type} æ•°æ®ä¸ºç©ºã€‚")
            else:
                # Aè‚¡æ•°æ®å·²ç»æ˜¯æ ¼å¼åŒ–åçš„ï¼ˆç§‘ç›®ã€ä¸­æ–‡ç§‘ç›®ã€æ•°å€¼(äº¿)ï¼‰ï¼Œæ¸¯è‚¡éœ€è¦è½¬ç½®
                if market == "Aè‚¡":
                    # Aè‚¡ï¼šç›´æ¥æ˜¾ç¤ºï¼Œä½†åªæ˜¾ç¤ºç§‘ç›®å’Œæ•°å€¼(äº¿)
                    display_df = df_preview[["ç§‘ç›®", "æ•°å€¼(äº¿)"]].copy()
                    display_df.columns = ["ç§‘ç›®", "æ•°å€¼(äº¿å…ƒ)"]
                else:
                    # æ¸¯è‚¡ï¼šè½¬ç½®ï¼Œå¹¶è½¬æ¢ä¸ºäº¿å…ƒï¼Œä½¿ç”¨ä¸­æ–‡åç§°
                    # æ’é™¤ REPORT_DATE åˆ—
                    df_to_transpose = df_preview.drop(columns=['REPORT_DATE'], errors='ignore')
                    df_t = df_to_transpose.T.reset_index()
                    df_t.columns = ["ç§‘ç›®", "æ•°å€¼"]
                    # å°†é‡‘é¢ä»å…ƒè½¬æ¢ä¸ºäº¿å…ƒï¼ˆè·³è¿‡éæ•°å€¼ï¼‰
                    def convert_to_yi(x):
                        if pd.isna(x) or str(x) == 'nan':
                            return x
                        try:
                            return round(float(x) / 100000000, 2)
                        except (ValueError, TypeError):
                            return x
                    df_t["æ•°å€¼"] = df_t["æ•°å€¼"].apply(convert_to_yi)
                    # ä½¿ç”¨ä¸­æ–‡åç§°æ˜ å°„æ›¿æ¢è‹±æ–‡å­—æ®µå
                    chinese_mappings = st.session_state.get(f"{session_key}_chinese_mappings", {})
                    stmt_key = stmt_map[sel_type]
                    chinese_mapping = chinese_mappings.get(stmt_key, {})
                    df_t["ç§‘ç›®"] = df_t["ç§‘ç›®"].apply(
                        lambda x: chinese_mapping.get(x, x)  # å¦‚æœæœ‰æ˜ å°„å°±ç”¨ä¸­æ–‡ï¼Œå¦åˆ™ç”¨åŸå€¼
                    )
                    display_df = df_t
                    display_df.columns = ["ç§‘ç›®", "æ•°å€¼(äº¿å…ƒ)"]
                # ç¡®ä¿"æ•°å€¼"åˆ—ä¸ºå­—ç¬¦ä¸²ç±»å‹ï¼Œé¿å… Arrow åºåˆ—åŒ–é”™è¯¯
                display_df["æ•°å€¼(äº¿å…ƒ)"] = display_df["æ•°å€¼(äº¿å…ƒ)"].astype(str)
                st.dataframe(display_df, width='stretch', height=420)
        
        # æ˜¾ç¤ºä¸‹è½½æŒ‰é’®
        if excel_bytes:
            filename = f"{company_name}_{start_year}-{end_year}_{'ä¸‰å¤§æŠ¥è¡¨' if market == 'Aè‚¡' else 'æ¸¯è‚¡ä¸‰å¤§æŠ¥è¡¨'}.xlsx"
            st.success("ä¸‹è½½å®Œæˆï¼Œå¯ä¿å­˜ä¸ºExcelã€‚")
            st.download_button("ğŸ“¥ ä¸‹è½½Excelæ–‡ä»¶", data=excel_bytes, file_name=filename,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        return

    try:
        if market == "Aè‚¡":
            # ä½¿ç”¨ç°æœ‰ä¸‹è½½å·¥å…·
            result_dict = dl_tool.get_financial_statements(symbol, start_year, end_year)
            if not result_dict:
                st.warning("æœªè·å–åˆ°ä»»ä½•æŠ¥è¡¨æ•°æ®ã€‚")
                return
            company_name = dl_tool.get_symbol_name(symbol)
            excel_bytes = dl_tool.create_excel_file(result_dict, symbol, company_name, start_year, end_year)
            
            # ä¿å­˜åˆ° session_state
            st.session_state[session_key] = result_dict
            st.session_state[f"{session_key}_excel"] = excel_bytes
            st.session_state[f"{session_key}_company"] = company_name
            
            # é‡æ–°è¿è¡Œä»¥æ˜¾ç¤ºé¢„è§ˆ
            st.rerun()
        else:
            # æ¸¯è‚¡ï¼šä½¿ç”¨é€‚é…å±‚è·å–ä¸‰å¤§æŠ¥è¡¨å¹¶ç”ŸæˆExcel
            data = hk_adapter.get_hk_annual_data(symbol, start_year, end_year)
            if not data or (data.get("profit") is None and data.get("balance_sheet") is None and data.get("cash_flow") is None):
                st.warning("æœªè·å–åˆ°æ¸¯è‚¡æŠ¥è¡¨æ•°æ®ã€‚")
                return

            def to_yearly(df, date_col="REPORT_DATE"):
                if df is None or df.empty:
                    return {}
                out = {}
                for year in range(start_year, end_year + 1):
                    row = hk_adapter.extract_year_data_hk(df, year, date_col_name=date_col)
                    if row is not None:
                        out[year] = row.to_frame().T
                return out

            profit_years = to_yearly(data.get("profit"))
            balance_years = to_yearly(data.get("balance_sheet"))
            cash_years = to_yearly(data.get("cash_flow"))
            
            # ä¿å­˜ä¸­æ–‡åç§°æ˜ å°„
            chinese_mappings = {
                "balance": data.get("balance_sheet_chinese_mapping", {}),
                "profit": data.get("profit_chinese_mapping", {}),
                "cash_flow": data.get("cash_flow_chinese_mapping", {}),
            }
            
            # ç»„è£…ä¸º {year: {balance, profit, cash}}
            results = {}
            for year in range(start_year, end_year + 1):
                year_data = {
                    "balance": balance_years.get(year),
                    "profit": profit_years.get(year),
                    "cash_flow": cash_years.get(year),
                }
                if any(v is not None for v in year_data.values()):
                    results[year] = year_data

            if not results:
                st.warning("æœªæ‰¾åˆ°æŒ‡å®šå¹´ä»½çš„æ¸¯è‚¡æŠ¥è¡¨æ•°æ®ã€‚")
                return

            # å†™å…¥Excelï¼ˆé‡‘é¢å·²è½¬æ¢ä¸ºäº¿å…ƒï¼‰
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                for year, year_data in results.items():
                    sheet_name = f"{year}å¹´"
                    pos = 0
                    for title, stmt_key, df in [
                        ("èµ„äº§è´Ÿå€ºè¡¨", "balance", year_data["balance"]),
                        ("åˆ©æ¶¦è¡¨", "profit", year_data["profit"]),
                        ("ç°é‡‘æµé‡è¡¨", "cash_flow", year_data["cash_flow"])
                    ]:
                        if df is None or df.empty:
                            continue
                        # è¡Œè½¬åˆ—ï¼šç§‘ç›® | æ•°å€¼ï¼ˆäº¿å…ƒï¼‰
                        # æ’é™¤ REPORT_DATE åˆ—
                        df_to_transpose = df.drop(columns=['REPORT_DATE'], errors='ignore')
                        df_t = df_to_transpose.T.reset_index()
                        df_t.columns = ["ç§‘ç›®", "æ•°å€¼"]
                        # å°†é‡‘é¢ä»å…ƒè½¬æ¢ä¸ºäº¿å…ƒï¼ˆè·³è¿‡éæ•°å€¼ï¼‰
                        def convert_to_yi(x):
                            if pd.isna(x) or str(x) == 'nan':
                                return x
                            try:
                                return round(float(x) / 100000000, 2)
                            except (ValueError, TypeError):
                                return x
                        df_t["æ•°å€¼"] = df_t["æ•°å€¼"].apply(convert_to_yi)
                        # ä½¿ç”¨ä¸­æ–‡åç§°æ˜ å°„æ›¿æ¢è‹±æ–‡å­—æ®µå
                        chinese_mapping = chinese_mappings.get(stmt_key, {})
                        df_t["ç§‘ç›®"] = df_t["ç§‘ç›®"].apply(
                            lambda x: chinese_mapping.get(x, x)  # å¦‚æœæœ‰æ˜ å°„å°±ç”¨ä¸­æ–‡ï¼Œå¦åˆ™ç”¨åŸå€¼
                        )
                        # å†™æ ‡é¢˜
                        ws = writer.book.create_sheet(sheet_name) if sheet_name not in writer.book.sheetnames else writer.book[sheet_name]
                        # å¦‚æœæ˜¯æ–°å»ºçš„ç©ºsheetï¼Œopenpyxlä¸ä¼šè¢«pandasè‡ªåŠ¨å†™å…¥ï¼›æ”¹ç”¨ pandas å†™æ•°æ®ï¼Œæ ‡é¢˜å•å…ƒæ ¼ç”¨ openpyxl
                        df_t.to_excel(writer, sheet_name=sheet_name, startrow=pos, index=False)
                        # å†™æ ‡é¢˜æ–‡å­—ï¼ˆæ”¾åœ¨æ•°æ®ä¸Šæ–¹ä¸€è¡Œï¼‰
                        ws = writer.sheets[sheet_name]
                        ws.cell(row=pos + 1, column=1, value=f"ã€{title}ã€‘")
                        pos += len(df_t) + 3  # æ ‡é¢˜ + æ•°æ® + ç©ºè¡Œ
                
                # åœ¨writerå…³é—­å‰è®¾ç½®åˆ—å®½è‡ªé€‚åº”
                try:
                    from openpyxl.utils import get_column_letter
                    
                    # ä¸ºæ¯ä¸ªsheetè®¾ç½®åˆ—å®½
                    for sheet_name in writer.book.sheetnames:
                        ws = writer.book[sheet_name]
                        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, values_only=False), start=1):
                            max_length = 0
                            column_letter = get_column_letter(col_idx)
                            
                            for cell in col:
                                if cell.value is not None:
                                    try:
                                        cell_value = str(cell.value)
                                        length = 0
                                        for char in cell_value:
                                            if ord(char) > 127:  # éASCIIå­—ç¬¦ï¼ˆåŒ…æ‹¬ä¸­æ–‡ï¼‰
                                                length += 2
                                            else:
                                                length += 1
                                        if length > max_length:
                                            max_length = length
                                    except:
                                        pass
                            
                            if max_length > 0:
                                adjusted_width = min(max(max_length + 2, 8), 50)
                                ws.column_dimensions[column_letter].width = adjusted_width
                            else:
                                ws.column_dimensions[column_letter].width = 10
                except Exception as e:
                    # å¦‚æœè®¾ç½®åˆ—å®½å¤±è´¥ï¼Œä¸å½±å“è¿”å›ç»“æœ
                    pass

            output.seek(0)
            excel_bytes = output.getvalue()
            
            # ä¿å­˜åˆ° session_stateï¼ˆåŒ…å«ä¸­æ–‡æ˜ å°„ï¼‰
            st.session_state[session_key] = results
            st.session_state[f"{session_key}_excel"] = excel_bytes
            st.session_state[f"{session_key}_company"] = symbol
            st.session_state[f"{session_key}_chinese_mappings"] = chinese_mappings
            
            # é‡æ–°è¿è¡Œä»¥æ˜¾ç¤ºé¢„è§ˆ
            st.rerun()

    except Exception as e:
        st.error(f"ä¸‹è½½å¤±è´¥ï¼š{e}")
        import traceback
        st.code(traceback.format_exc())

# -----------------------------
# åŠŸèƒ½ 3ï¼šå‘˜å·¥æ•°é‡æå–
# -----------------------------
def run_employee_extraction():
    st.header("ğŸ‘¥ å‘˜å·¥æ•°é‡æå–")
    
    # æ–‡ä»¶å¤¹é€‰æ‹©åŠŸèƒ½
    def select_folder():
        """æ‰“å¼€æ–‡ä»¶å¤¹é€‰æ‹©å¯¹è¯æ¡†"""
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()  # éšè—ä¸»çª—å£
            root.attributes('-topmost', True)  # çª—å£ç½®é¡¶
            folder_path = filedialog.askdirectory(title="é€‰æ‹©PDFæ–‡ä»¶å¤¹")
            root.destroy()
            return folder_path if folder_path else None
        except Exception as e:
            st.warning(f"æ–‡ä»¶å¤¹é€‰æ‹©å™¨ä¸å¯ç”¨: {e}ï¼Œè¯·æ‰‹åŠ¨è¾“å…¥è·¯å¾„")
            return None
    
    # åˆå§‹åŒ–PDFç›®å½•è·¯å¾„
    if 'pdf_dir' not in st.session_state:
        st.session_state['pdf_dir'] = "å¹´æŠ¥PDF"
    
    # æ–‡ä»¶å¤¹é€‰æ‹©æŒ‰é’®å’Œè¾“å…¥æ¡†
    col1, col2 = st.sidebar.columns([3, 1])
    with col1:
        # ä¸ä½¿ç”¨keyå‚æ•°ï¼Œç›´æ¥é€šè¿‡valueå‚æ•°æ§åˆ¶ï¼Œè¿™æ ·é€‰æ‹©æ–‡ä»¶å¤¹åå¯ä»¥ç«‹å³æ›´æ–°
        pdf_dir = st.text_input(
            "PDFç›®å½•è·¯å¾„", 
            value=st.session_state.get('pdf_dir', "å¹´æŠ¥PDF")
        )
        # å¦‚æœè¾“å…¥æ¡†çš„å€¼æ”¹å˜äº†ï¼Œæ›´æ–°session_state
        if pdf_dir:
            st.session_state['pdf_dir'] = pdf_dir
    with col2:
        if st.button("ğŸ“ é€‰æ‹©", use_container_width=True, help="ç‚¹å‡»é€‰æ‹©æ–‡ä»¶å¤¹", key="select_folder_btn"):
            selected_folder = select_folder()
            if selected_folder:
                # ç›´æ¥æ›´æ–°session_stateï¼Œç„¶årerunï¼Œè¾“å…¥æ¡†ä¼šè‡ªåŠ¨ä½¿ç”¨æ–°çš„value
                st.session_state['pdf_dir'] = selected_folder
                st.rerun()  # åˆ·æ–°é¡µé¢ä»¥æ›´æ–°è¾“å…¥æ¡†
    
    run_btn = st.sidebar.button("ğŸš€ å¼€å§‹æå–", type="primary", use_container_width=True)
    if not run_btn:
        st.info("é€‰æ‹©æˆ–è¾“å…¥PDFç›®å½•å¹¶ç‚¹å‡»å¼€å§‹æå–ã€‚")
        return

    try:
        # ä½¿ç”¨session_stateä¸­çš„pdf_dir
        actual_pdf_dir = st.session_state.get('pdf_dir', pdf_dir)
        
        # æ˜¾ç¤ºå½“å‰ä½¿ç”¨çš„è·¯å¾„ï¼ˆç”¨äºè°ƒè¯•ï¼‰
        if actual_pdf_dir:
            st.info(f"ğŸ“‚ å½“å‰PDFç›®å½•: `{actual_pdf_dir}`")
        
        if not actual_pdf_dir:
            st.error("âŒ è¯·é€‰æ‹©æˆ–è¾“å…¥PDFç›®å½•è·¯å¾„")
            return
            
        # æ£€æŸ¥ç›®å½•æ˜¯å¦å­˜åœ¨
        if not os.path.exists(actual_pdf_dir):
            st.error(f"âŒ PDFç›®å½•ä¸å­˜åœ¨: `{actual_pdf_dir}`")
            st.info("ğŸ’¡ è¯·æ£€æŸ¥è·¯å¾„æ˜¯å¦æ­£ç¡®ï¼Œæˆ–ä½¿ç”¨ã€ğŸ“ é€‰æ‹©ã€‘æŒ‰é’®é‡æ–°é€‰æ‹©æ–‡ä»¶å¤¹")
            return
        
        # æ£€æŸ¥ç›®å½•ä¸­æ˜¯å¦æœ‰PDFæ–‡ä»¶
        pdf_files = [f for f in os.listdir(actual_pdf_dir) if f.lower().endswith('.pdf')]
        if not pdf_files:
            st.warning(f"âš ï¸ ç›®å½•ä¸­æ²¡æœ‰æ‰¾åˆ°PDFæ–‡ä»¶: `{actual_pdf_dir}`")
            st.info("ğŸ’¡ è¯·ç¡®ä¿ç›®å½•ä¸­åŒ…å«å¹´æŠ¥PDFæ–‡ä»¶")
        else:
            st.success(f"âœ“ æ‰¾åˆ° {len(pdf_files)} ä¸ªPDFæ–‡ä»¶")
        
        results = {}
        if market == "Aè‚¡":
            # æ‰¹é‡æå–ï¼ˆä¼ é€’è‚¡ç¥¨ä»£ç ï¼‰
            res = emp_a.batch_extract_employee_count_smart(actual_pdf_dir, stock_code=symbol, use_smart=True)
            results = {k: v for k, v in res.items()}
        else:
            # æ¸¯è‚¡æŒ‰å¹´ä»½æå–
            res = emp_hk.extract_employee_count_by_year_from_pdfs(actual_pdf_dir, symbol, start_year, end_year)
            results = {f"{year}å¹´": count for year, count in res.items()}

        if not results:
            st.warning("æœªæå–åˆ°å‘˜å·¥æ•°é‡ï¼Œè¯·æ£€æŸ¥PDFç›®å½•ã€‚")
            return

        # å±•ç¤ºç»“æœ
        df = pd.DataFrame(list(results.items()), columns=["å¹´ä»½/æ–‡ä»¶", "å‘˜å·¥æ•°é‡"])
        st.dataframe(df, width='stretch', height=400)

        # ä¸‹è½½CSV
        output = io.StringIO()
        df.to_csv(output, index=False, encoding="utf-8-sig")
        st.download_button("ğŸ“¥ ä¸‹è½½å‘˜å·¥æ•°é‡CSV", data=output.getvalue().encode("utf-8-sig"),
                           file_name=f"{symbol}_å‘˜å·¥æ•°é‡.csv", mime="text/csv")

    except Exception as e:
        st.error(f"æå–å¤±è´¥ï¼š{e}")
        import traceback
        st.code(traceback.format_exc())

# -----------------------------
# åŠŸèƒ½ 4ï¼šå¹´æŠ¥PDFä¸‹è½½ï¼ˆAè‚¡ + æ¸¯è‚¡ï¼‰
# -----------------------------
def run_pdf_download():
    if market == "Aè‚¡":
        run_pdf_download_a()
    else:
        run_pdf_download_hk()


def run_pdf_download_a():
    """Aè‚¡å¹´æŠ¥PDFä¸‹è½½"""
    st.header("ğŸ“¥ å¹´æŠ¥PDFä¸‹è½½ï¼ˆAè‚¡ï¼‰")
    st.info("ä»å·¨æ½®èµ„è®¯ç½‘ä¸‹è½½Aè‚¡ä¸Šå¸‚å…¬å¸å¹´åº¦æŠ¥å‘ŠPDF")
    
    # åŠ è½½ä¸‹è½½æ¨¡å—
    try:
        pdf_dl = load_module("pdf_downloader", "08_ä¸‹è½½å¹´æŠ¥PDF.py")
    except Exception as e:
        st.error(f"åŠ è½½ä¸‹è½½æ¨¡å—å¤±è´¥ï¼š{e}")
        return
    
    # æ–‡ä»¶å¤¹é€‰æ‹©åŠŸèƒ½
    def select_save_folder():
        """æ‰“å¼€æ–‡ä»¶å¤¹é€‰æ‹©å¯¹è¯æ¡†"""
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            folder_path = filedialog.askdirectory(title="é€‰æ‹©PDFä¿å­˜ç›®å½•")
            root.destroy()
            return folder_path if folder_path else None
        except Exception as e:
            st.warning(f"æ–‡ä»¶å¤¹é€‰æ‹©å™¨ä¸å¯ç”¨: {e}ï¼Œè¯·æ‰‹åŠ¨è¾“å…¥è·¯å¾„")
            return None
    
    # åˆå§‹åŒ–ä¿å­˜è·¯å¾„
    if 'pdf_save_dir' not in st.session_state:
        st.session_state['pdf_save_dir'] = "å¹´æŠ¥PDF"
    
    st.sidebar.markdown("---")
    st.sidebar.subheader("ğŸ“ ä¸‹è½½è®¾ç½®")
    
    # ä¿å­˜è·¯å¾„é€‰æ‹©ï¼ˆä¸ä½¿ç”¨keyå‚æ•°ï¼Œé€šè¿‡valueç›´æ¥æ§åˆ¶ï¼‰
    col1, col2 = st.sidebar.columns([3, 1])
    with col1:
        save_dir = st.text_input(
            "ä¿å­˜ç›®å½•",
            value=st.session_state.get('pdf_save_dir', "å¹´æŠ¥PDF")
        )
        # å¦‚æœç”¨æˆ·æ‰‹åŠ¨è¾“å…¥äº†è·¯å¾„ï¼Œæ›´æ–°session_state
        if save_dir and save_dir != st.session_state.get('pdf_save_dir'):
            st.session_state['pdf_save_dir'] = save_dir
    with col2:
        if st.button("ğŸ“", use_container_width=True, help="é€‰æ‹©ä¿å­˜æ–‡ä»¶å¤¹", key="select_save_folder_btn"):
            selected_folder = select_save_folder()
            if selected_folder:
                st.session_state['pdf_save_dir'] = selected_folder
                st.rerun()
    
    # ä¸‹è½½æŒ‰é’®
    download_btn = st.sidebar.button("ğŸš€ å¼€å§‹ä¸‹è½½", type="primary", use_container_width=True, key="download_pdf_btn")
    
    # æ˜¾ç¤ºå½“å‰è®¾ç½®
    st.markdown("### ğŸ“‹ ä¸‹è½½è®¾ç½®")
    col_info1, col_info2 = st.columns(2)
    with col_info1:
        st.write(f"**è‚¡ç¥¨ä»£ç ï¼š** {symbol}")
        st.write(f"**å¹´ä»½èŒƒå›´ï¼š** {start_year} - {end_year}")
    with col_info2:
        actual_save_dir = st.session_state.get('pdf_save_dir', save_dir)
        st.write(f"**ä¿å­˜ç›®å½•ï¼š** `{actual_save_dir}`")
        # æ£€æŸ¥ç›®å½•
        if os.path.exists(actual_save_dir):
            existing_pdfs = [f for f in os.listdir(actual_save_dir) if f.lower().endswith('.pdf')]
            st.write(f"**å·²æœ‰æ–‡ä»¶ï¼š** {len(existing_pdfs)} ä¸ªPDF")
        else:
            st.write("**ç›®å½•çŠ¶æ€ï¼š** å°†è‡ªåŠ¨åˆ›å»º")
    
    if not download_btn:
        st.markdown("---")
        st.markdown("""
        ### ğŸ“– ä½¿ç”¨è¯´æ˜
        1. åœ¨å·¦ä¾§è¾“å…¥ **è‚¡ç¥¨ä»£ç **ï¼ˆ6ä½æ•°å­—ï¼Œå¦‚ 600900ï¼‰
        2. è®¾ç½® **èµ·å§‹å¹´ä»½** å’Œ **ç»“æŸå¹´ä»½**
        3. é€‰æ‹©æˆ–è¾“å…¥ **PDFä¿å­˜ç›®å½•**
        4. ç‚¹å‡» **å¼€å§‹ä¸‹è½½** æŒ‰é’®
        
        ### âš ï¸ æ³¨æ„äº‹é¡¹
        - å¹´æŠ¥PDFé€šå¸¸åœ¨æ¬¡å¹´3-4æœˆå‘å¸ƒï¼ˆå¦‚2023å¹´å¹´æŠ¥åœ¨2024å¹´4æœˆå‰å‘å¸ƒï¼‰
        - ä¸‹è½½éœ€è¦ç½‘ç»œè¿æ¥ï¼Œè¯·ç¡®ä¿ç½‘ç»œé€šç•…
        - å•ä¸ªå¹´æŠ¥PDFæ–‡ä»¶è¾ƒå¤§ï¼ˆå‡ MBåˆ°å‡ åMBï¼‰ï¼Œè¯·è€å¿ƒç­‰å¾…
        """)
        return
    
    # æ‰§è¡Œä¸‹è½½
    st.markdown("---")
    st.markdown("### ğŸ“¥ ä¸‹è½½è¿›åº¦")
    
    actual_save_dir = st.session_state.get('pdf_save_dir', save_dir)
    
    # åˆ›å»ºä¿å­˜ç›®å½•
    os.makedirs(actual_save_dir, exist_ok=True)
    
    # ä¸‹è½½ç»“æœç»Ÿè®¡
    results = {
        'success': [],
        'failed': []
    }
    
    # è¿›åº¦æ¡
    progress_bar = st.progress(0)
    status_text = st.empty()
    log_container = st.container()
    
    years = list(range(int(start_year), int(end_year) + 1))
    total_years = len(years)
    
    for idx, year in enumerate(years):
        status_text.text(f"æ­£åœ¨ä¸‹è½½ {year} å¹´å¹´æŠ¥... ({idx + 1}/{total_years})")
        
        with log_container:
            st.write(f"**[{year}å¹´]** æœç´¢ä¸­...")
        
        try:
            # è°ƒç”¨ä¸‹è½½å‡½æ•°
            filepath = pdf_dl.download_annual_report(symbol, year, actual_save_dir)
            
            if filepath and os.path.exists(filepath):
                file_size = os.path.getsize(filepath) / 1024 / 1024  # MB
                results['success'].append({
                    'year': year,
                    'path': filepath,
                    'size': f"{file_size:.2f} MB"
                })
                with log_container:
                    st.success(f"[OK] {year}å¹´å¹´æŠ¥ä¸‹è½½æˆåŠŸï¼š{os.path.basename(filepath)} ({file_size:.2f} MB)")
            else:
                results['failed'].append({
                    'year': year,
                    'reason': 'æœªæ‰¾åˆ°å¹´æŠ¥æˆ–ä¸‹è½½å¤±è´¥'
                })
                with log_container:
                    st.warning(f"[!] {year}å¹´å¹´æŠ¥ä¸‹è½½å¤±è´¥")
        except Exception as e:
            results['failed'].append({
                'year': year,
                'reason': str(e)
            })
            with log_container:
                st.error(f"[X] {year}å¹´å¹´æŠ¥ä¸‹è½½å‡ºé”™ï¼š{e}")
        
        # æ›´æ–°è¿›åº¦
        progress_bar.progress((idx + 1) / total_years)
    
    # æ˜¾ç¤ºä¸‹è½½ç»“æœæ±‡æ€»
    st.markdown("---")
    st.markdown("### ğŸ“Š ä¸‹è½½ç»“æœ")
    
    col_success, col_failed = st.columns(2)
    
    with col_success:
        st.metric("ä¸‹è½½æˆåŠŸ", f"{len(results['success'])} ä¸ª")
        if results['success']:
            for item in results['success']:
                st.write(f"- {item['year']}å¹´ï¼š{item['size']}")
    
    with col_failed:
        st.metric("ä¸‹è½½å¤±è´¥", f"{len(results['failed'])} ä¸ª")
        if results['failed']:
            for item in results['failed']:
                st.write(f"- {item['year']}å¹´ï¼š{item['reason']}")
    
    # æ‰“å¼€ä¿å­˜ç›®å½•æŒ‰é’®
    if results['success']:
        st.success(f"ä¸‹è½½å®Œæˆï¼æ–‡ä»¶ä¿å­˜åœ¨ï¼š`{actual_save_dir}`")


def run_pdf_download_hk():
    """æ¸¯è‚¡å¹´æŠ¥PDFä¸‹è½½ï¼ˆä»HTMLæ–‡ä»¶è§£æï¼‰"""
    st.header("ğŸ“¥ å¹´æŠ¥PDFä¸‹è½½ï¼ˆæ¸¯è‚¡ï¼‰")
    st.info("ä»æ¸¯äº¤æ‰€æŠ«éœ²æ˜“ä¸‹è½½æ¸¯è‚¡å¹´åº¦æŠ¥å‘ŠPDFï¼ˆéœ€å…ˆä¿å­˜æœç´¢ç»“æœHTMLï¼‰")
    
    # åŠ è½½æ¸¯è‚¡ä¸‹è½½æ¨¡å—
    try:
        hk_pdf_dl = load_module("hk_pdf_downloader", "09_ä¸‹è½½æ¸¯è‚¡å¹´æŠ¥PDF.py")
    except Exception as e:
        st.error(f"åŠ è½½æ¸¯è‚¡ä¸‹è½½æ¨¡å—å¤±è´¥ï¼š{e}")
        return
    
    # æ–‡ä»¶é€‰æ‹©åŠŸèƒ½
    def select_html_file():
        """æ‰“å¼€æ–‡ä»¶é€‰æ‹©å¯¹è¯æ¡†é€‰æ‹©HTMLæ–‡ä»¶"""
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            file_path = filedialog.askopenfilename(
                title="é€‰æ‹©æ¸¯äº¤æ‰€æœç´¢ç»“æœHTMLæ–‡ä»¶",
                filetypes=[("HTMLæ–‡ä»¶", "*.html;*.htm"), ("æ‰€æœ‰æ–‡ä»¶", "*.*")]
            )
            root.destroy()
            return file_path if file_path else None
        except Exception as e:
            st.warning(f"æ–‡ä»¶é€‰æ‹©å™¨ä¸å¯ç”¨: {e}ï¼Œè¯·æ‰‹åŠ¨è¾“å…¥è·¯å¾„")
            return None
    
    def select_save_folder():
        """æ‰“å¼€æ–‡ä»¶å¤¹é€‰æ‹©å¯¹è¯æ¡†"""
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            folder_path = filedialog.askdirectory(title="é€‰æ‹©PDFä¿å­˜ç›®å½•")
            root.destroy()
            return folder_path if folder_path else None
        except Exception as e:
            st.warning(f"æ–‡ä»¶å¤¹é€‰æ‹©å™¨ä¸å¯ç”¨: {e}ï¼Œè¯·æ‰‹åŠ¨è¾“å…¥è·¯å¾„")
            return None
    
    # åˆå§‹åŒ–è·¯å¾„
    if 'hk_html_path' not in st.session_state:
        st.session_state['hk_html_path'] = ""
    if 'hk_pdf_save_dir' not in st.session_state:
        st.session_state['hk_pdf_save_dir'] = "æ¸¯è‚¡å¹´æŠ¥PDF"
    
    st.sidebar.markdown("---")
    st.sidebar.subheader("ğŸ“ ä¸‹è½½è®¾ç½®")
    
    # HTMLæ–‡ä»¶é€‰æ‹©
    st.sidebar.markdown("**æ­¥éª¤1: é€‰æ‹©HTMLæ–‡ä»¶**")
    col1, col2 = st.sidebar.columns([3, 1])
    with col1:
        html_path = st.text_input(
            "HTMLæ–‡ä»¶è·¯å¾„",
            value=st.session_state.get('hk_html_path', ""),
            placeholder="ä»æ¸¯äº¤æ‰€ä¿å­˜çš„æœç´¢ç»“æœHTML"
        )
        # å¦‚æœç”¨æˆ·æ‰‹åŠ¨è¾“å…¥äº†è·¯å¾„ï¼Œæ›´æ–°session_state
        if html_path and html_path != st.session_state.get('hk_html_path'):
            st.session_state['hk_html_path'] = html_path
    with col2:
        if st.button("ğŸ“„", use_container_width=True, help="é€‰æ‹©HTMLæ–‡ä»¶", key="select_html_btn"):
            selected_file = select_html_file()
            if selected_file:
                st.session_state['hk_html_path'] = selected_file
                st.rerun()
    
    # PDFä¿å­˜ç›®å½•é€‰æ‹©
    st.sidebar.markdown("**æ­¥éª¤2: é€‰æ‹©ä¿å­˜ç›®å½•**")
    col3, col4 = st.sidebar.columns([3, 1])
    with col3:
        save_dir = st.text_input(
            "ä¿å­˜ç›®å½•",
            value=st.session_state.get('hk_pdf_save_dir', "æ¸¯è‚¡å¹´æŠ¥PDF")
        )
        # å¦‚æœç”¨æˆ·æ‰‹åŠ¨è¾“å…¥äº†è·¯å¾„ï¼Œæ›´æ–°session_state
        if save_dir and save_dir != st.session_state.get('hk_pdf_save_dir'):
            st.session_state['hk_pdf_save_dir'] = save_dir
    with col4:
        if st.button("ğŸ“", use_container_width=True, help="é€‰æ‹©ä¿å­˜æ–‡ä»¶å¤¹", key="select_hk_save_folder_btn"):
            selected_folder = select_save_folder()
            if selected_folder:
                st.session_state['hk_pdf_save_dir'] = selected_folder
                st.rerun()
    
    # ä¸‹è½½æŒ‰é’®
    download_btn = st.sidebar.button("ğŸš€ å¼€å§‹ä¸‹è½½", type="primary", use_container_width=True, key="download_hk_pdf_btn")
    
    # æ˜¾ç¤ºå½“å‰è®¾ç½®
    st.markdown("### ğŸ“‹ ä¸‹è½½è®¾ç½®")
    col_info1, col_info2 = st.columns(2)
    with col_info1:
        st.write(f"**è‚¡ç¥¨ä»£ç ï¼š** {symbol}")
        st.write(f"**å¹´ä»½èŒƒå›´ï¼š** {start_year} - {end_year}")
    with col_info2:
        actual_html_path = st.session_state.get('hk_html_path', '')
        actual_save_dir = st.session_state.get('hk_pdf_save_dir', save_dir)
        if actual_html_path:
            st.write(f"**HTMLæ–‡ä»¶ï¼š** `{os.path.basename(actual_html_path)}`")
        else:
            st.write("**HTMLæ–‡ä»¶ï¼š** æœªé€‰æ‹©")
        st.write(f"**ä¿å­˜ç›®å½•ï¼š** `{actual_save_dir}`")
    
    # å¦‚æœæ²¡æœ‰ç‚¹å‡»ä¸‹è½½æŒ‰é’®ï¼Œæ˜¾ç¤ºä½¿ç”¨è¯´æ˜
    if not download_btn:
        st.markdown("---")
        st.markdown("""
        ### ğŸ“– ä½¿ç”¨è¯´æ˜
        
        **æ¸¯è‚¡å¹´æŠ¥ä¸‹è½½éœ€è¦å…ˆä»æ¸¯äº¤æ‰€ç½‘ç«™ä¿å­˜æœç´¢ç»“æœé¡µé¢ï¼š**
        
        1. æ‰“å¼€æ¸¯äº¤æ‰€æŠ«éœ²æ˜“æœç´¢é¡µé¢ï¼š[https://www1.hkexnews.hk/search/titlesearch.xhtml](https://www1.hkexnews.hk/search/titlesearch.xhtml)
        2. è¾“å…¥è‚¡ç¥¨ä»£ç ï¼ˆå¦‚ 01810ï¼‰ï¼Œé€‰æ‹© **"å¹´åº¦æŠ¥å‘Š"** æ–‡ä»¶ç±»åˆ«
        3. ç‚¹å‡»æœç´¢ï¼Œç­‰å¾…ç»“æœæ˜¾ç¤º
        4. **Ctrl+S** ä¿å­˜ç½‘é¡µä¸ºHTMLæ–‡ä»¶ï¼ˆå®Œæ•´ç½‘é¡µæ ¼å¼ï¼‰
        5. å›åˆ°æœ¬å·¥å…·ï¼Œé€‰æ‹©åˆšä¿å­˜çš„HTMLæ–‡ä»¶
        6. è®¾ç½®PDFä¿å­˜ç›®å½•
        7. ç‚¹å‡» **å¼€å§‹ä¸‹è½½**
        
        ### âš ï¸ æ³¨æ„äº‹é¡¹
        - ä¿å­˜HTMLæ—¶è¯·é€‰æ‹© **"ç½‘é¡µï¼Œå®Œæ•´"** æˆ– **"ç½‘é¡µï¼Œä»…HTML"** æ ¼å¼
        - å¹´ä»½èŒƒå›´ä¼šç”¨äºç­›é€‰è¦ä¸‹è½½çš„å¹´æŠ¥
        - ä¸‹è½½é€Ÿåº¦å–å†³äºç½‘ç»œçŠ¶å†µ
        """)
        
        # æ˜¾ç¤ºè§£æé¢„è§ˆï¼ˆå¦‚æœå·²é€‰æ‹©HTMLæ–‡ä»¶ï¼‰
        if actual_html_path and os.path.exists(actual_html_path):
            st.markdown("---")
            st.markdown("### ğŸ” HTMLæ–‡ä»¶é¢„è§ˆ")
            try:
                reports = hk_pdf_dl.parse_html_for_annual_reports(actual_html_path)
                if reports:
                    # ç­›é€‰å¹´ä»½èŒƒå›´
                    filtered_reports = [r for r in reports if start_year <= r['year'] <= end_year]
                    st.success(f"è§£ææˆåŠŸï¼æ‰¾åˆ° {len(reports)} ä¸ªå¹´æŠ¥ï¼Œç¬¦åˆå¹´ä»½èŒƒå›´çš„æœ‰ {len(filtered_reports)} ä¸ª")
                    
                    # æ˜¾ç¤ºåˆ—è¡¨
                    preview_data = []
                    for r in reports:
                        in_range = "âœ“" if start_year <= r['year'] <= end_year else ""
                        preview_data.append({
                            "é€‰ä¸­": in_range,
                            "å¹´ä»½": r['year'],
                            "æ ‡é¢˜": r['title'][:40],
                        })
                    st.dataframe(pd.DataFrame(preview_data), use_container_width=True)
                else:
                    st.warning("æœªåœ¨HTMLæ–‡ä»¶ä¸­æ‰¾åˆ°å¹´æŠ¥é“¾æ¥ï¼Œè¯·æ£€æŸ¥æ–‡ä»¶æ˜¯å¦æ­£ç¡®")
            except Exception as e:
                st.error(f"è§£æHTMLå¤±è´¥ï¼š{e}")
        return
    
    # æ‰§è¡Œä¸‹è½½
    actual_html_path = st.session_state.get('hk_html_path', '')
    actual_save_dir = st.session_state.get('hk_pdf_save_dir', save_dir)
    
    # æ£€æŸ¥HTMLæ–‡ä»¶
    if not actual_html_path:
        st.error("è¯·å…ˆé€‰æ‹©HTMLæ–‡ä»¶ï¼")
        return
    if not os.path.exists(actual_html_path):
        st.error(f"HTMLæ–‡ä»¶ä¸å­˜åœ¨ï¼š{actual_html_path}")
        return
    
    # åˆ›å»ºä¿å­˜ç›®å½•
    os.makedirs(actual_save_dir, exist_ok=True)
    
    st.markdown("---")
    st.markdown("### ğŸ“¥ ä¸‹è½½è¿›åº¦")
    
    # è§£æHTMLè·å–å¹´æŠ¥åˆ—è¡¨
    status_text = st.empty()
    status_text.text("æ­£åœ¨è§£æHTMLæ–‡ä»¶...")
    
    try:
        reports = hk_pdf_dl.parse_html_for_annual_reports(actual_html_path)
        if not reports:
            st.error("æœªåœ¨HTMLæ–‡ä»¶ä¸­æ‰¾åˆ°å¹´æŠ¥é“¾æ¥")
            return
        
        # ç­›é€‰å¹´ä»½èŒƒå›´
        years_to_download = list(range(int(start_year), int(end_year) + 1))
        filtered_reports = [r for r in reports if r['year'] in years_to_download]
        
        if not filtered_reports:
            st.warning(f"æ²¡æœ‰æ‰¾åˆ° {start_year}-{end_year} å¹´ä»½èŒƒå›´å†…çš„å¹´æŠ¥")
            st.info(f"HTMLæ–‡ä»¶ä¸­åŒ…å«çš„å¹´ä»½ï¼š{sorted([r['year'] for r in reports])}")
            return
        
        st.info(f"å‡†å¤‡ä¸‹è½½ {len(filtered_reports)} ä¸ªå¹´æŠ¥")
        
        # ä¸‹è½½ç»“æœç»Ÿè®¡
        results = {'success': [], 'failed': []}
        
        # è¿›åº¦æ¡
        progress_bar = st.progress(0)
        log_container = st.container()
        
        total = len(filtered_reports)
        for idx, report in enumerate(filtered_reports):
            year = report['year']
            pdf_url = report['pdf_url']
            title = report['title']
            
            status_text.text(f"æ­£åœ¨ä¸‹è½½ {year} å¹´å¹´æŠ¥... ({idx + 1}/{total})")
            
            with log_container:
                st.write(f"**[{year}å¹´]** {title[:30]}...")
            
            try:
                # ç”Ÿæˆæ–‡ä»¶å
                symbol_clean = symbol.zfill(5)
                filename = f"{symbol_clean}_{year}å¹´å¹´åº¦æŠ¥å‘Š.pdf"
                save_path = os.path.join(actual_save_dir, filename)
                
                # ä¸‹è½½PDF
                success = hk_pdf_dl.download_pdf_from_url(pdf_url, save_path)
                
                if success and os.path.exists(save_path):
                    file_size = os.path.getsize(save_path) / 1024 / 1024
                    results['success'].append({
                        'year': year,
                        'path': save_path,
                        'size': f"{file_size:.2f} MB"
                    })
                    with log_container:
                        st.success(f"[OK] {year}å¹´å¹´æŠ¥ä¸‹è½½æˆåŠŸ ({file_size:.2f} MB)")
                else:
                    results['failed'].append({
                        'year': year,
                        'reason': 'ä¸‹è½½å¤±è´¥'
                    })
                    with log_container:
                        st.warning(f"[!] {year}å¹´å¹´æŠ¥ä¸‹è½½å¤±è´¥")
            except Exception as e:
                results['failed'].append({
                    'year': year,
                    'reason': str(e)
                })
                with log_container:
                    st.error(f"[X] {year}å¹´å¹´æŠ¥ä¸‹è½½å‡ºé”™ï¼š{e}")
            
            # æ›´æ–°è¿›åº¦
            progress_bar.progress((idx + 1) / total)
        
        # æ˜¾ç¤ºä¸‹è½½ç»“æœæ±‡æ€»
        st.markdown("---")
        st.markdown("### ğŸ“Š ä¸‹è½½ç»“æœ")
        
        col_success, col_failed = st.columns(2)
        
        with col_success:
            st.metric("ä¸‹è½½æˆåŠŸ", f"{len(results['success'])} ä¸ª")
            if results['success']:
                for item in results['success']:
                    st.write(f"- {item['year']}å¹´ï¼š{item['size']}")
        
        with col_failed:
            st.metric("ä¸‹è½½å¤±è´¥", f"{len(results['failed'])} ä¸ª")
            if results['failed']:
                for item in results['failed']:
                    st.write(f"- {item['year']}å¹´ï¼š{item['reason']}")
        
        if results['success']:
            st.success(f"ä¸‹è½½å®Œæˆï¼æ–‡ä»¶ä¿å­˜åœ¨ï¼š`{actual_save_dir}`")
    
    except Exception as e:
        st.error(f"ä¸‹è½½å¤±è´¥ï¼š{e}")
        import traceback
        st.code(traceback.format_exc())


# -----------------------------
# ä¸»è·¯ç”±
# -----------------------------
st.title("ğŸ“Š ç»Ÿä¸€è´¢åŠ¡å·¥å…·ï¼ˆAè‚¡ + æ¸¯è‚¡ï¼‰")
st.caption("è´¢åŠ¡åˆ†æï½œæŠ¥è¡¨ä¸‹è½½ï½œå‘˜å·¥æ•°é‡æå–ï½œå¹´æŠ¥ä¸‹è½½ â€”â€” ä¸€ç«™å¼ç•Œé¢")
st.markdown("---")

if feature == "ğŸ“Š è´¢åŠ¡åˆ†æ":
    run_financial_analysis()
elif feature == "ğŸ“„ æŠ¥è¡¨ä¸‹è½½":
    run_report_download()
elif feature == "ğŸ‘¥ å‘˜å·¥æ•°é‡æå–":
    run_employee_extraction()
elif feature == "ğŸ“¥ å¹´æŠ¥PDFä¸‹è½½":
    run_pdf_download()
else:
    st.info("è¯·é€‰æ‹©å·¦ä¾§çš„åŠŸèƒ½å¼€å§‹ã€‚")

