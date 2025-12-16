"""
财务报表下载工具 - Streamlit界面

功能：
1. 输入股票代码和起止年份
2. 获取历年三大财务报表数据（资产负债表、利润表、现金流量表）
3. 按年份组织，每年一个sheet
4. 每个sheet从上到下显示：资产负债表、利润表、现金流量表
5. 支持下载Excel文件
"""

import streamlit as st
import akshare as ak
import pandas as pd
import os
from datetime import datetime
from typing import Optional
import io

# 导入06_格式化显示财务数据模块
import importlib.util
import sys

# 动态导入06_格式化显示财务数据模块
module_name = "06_格式化显示财务数据"
module_path = "06_格式化显示财务数据.py"
spec = importlib.util.spec_from_file_location(module_name, module_path)
format_module = importlib.util.module_from_spec(spec)
sys.modules[module_name] = format_module
spec.loader.exec_module(format_module)

# 从模块中获取需要的函数和映射
FINANCIAL_ITEM_MAPPING = format_module.FINANCIAL_ITEM_MAPPING
get_chinese_name = format_module.get_chinese_name
convert_to_yi = format_module.convert_to_yi

def get_symbol_with_suffix(symbol):
    """
    为股票代码添加交易所后缀
    
    参数:
        symbol: 股票代码（如 "000001" 或 "600519"）
    
    返回:
        带后缀的股票代码（如 "000001.SZ" 或 "600519.SH"）
    """
    if '.' in symbol:
        return symbol
    
    if symbol.startswith(('000', '001', '002', '300')):
        return symbol + '.SZ'
    elif symbol.startswith(('600', '601', '603', '605', '688')):
        return symbol + '.SH'
    else:
        return symbol + '.SZ'

def get_symbol_name(symbol):
    """
    获取股票名称
    
    参数:
        symbol: 股票代码
    
    返回:
        股票名称
    """
    try:
        symbol_clean = symbol.replace('.SZ', '').replace('.SH', '')
        stock_info = ak.stock_individual_info_em(symbol=symbol_clean)
        if stock_info is not None and not stock_info.empty:
            name_row = stock_info[stock_info['item'] == '股票简称']
            if not name_row.empty:
                return name_row.iloc[0]['value']
    except:
        pass
    return symbol.replace('.SZ', '').replace('.SH', '')

def format_statement_data(df, year, statement_type: Optional[str] = None):
    """
    格式化财务报表数据为"每个科目一行"的格式
    
    参数:
        df: 原始数据框
        year: 年份
    
    返回:
        格式化后的数据框，包含：科目、中文科目、数值(亿)
    """
    if df is None or df.empty:
        return None
    
    # 查找日期列
    date_col = None
    for col in df.columns:
        if 'REPORT_DATE' in col or '报告期' in col:
            date_col = col
            break
    
    if date_col is None:
        return None
    
    # 筛选指定年份的数据（12-31年报）
    year_str = str(year)
    date_str = f"{year}-12-31"
    date_str_alt = f"{year_str}-12-31"
    filtered = df[
        df[date_col].astype(str).str.contains(date_str, na=False) |
        df[date_col].astype(str).str.contains(date_str_alt, na=False)
    ]
    
    if filtered.empty:
        return None
    
    # 取第一行（如果有多行）
    row_data = filtered.iloc[0]
    
    # 转置：每列变成一行
    result_data = []
    exclude_cols = [date_col, 'SECUCODE', 'SECURITY_CODE', 'SECURITY_NAME_ABBR', 
                  'ORG_CODE', 'ORG_TYPE', 'REPORT_TYPE', 'REPORT_DATE_NAME',
                  'SECURITY_TYPE_CODE', 'NOTICE_DATE', 'UPDATE_DATE', 'CURRENCY',
                  'OPINION_TYPE', 'OSOPINION_TYPE', 'LISTING_STATE']
    
    for col in df.columns:
        if col not in exclude_cols:
            value = row_data[col]
            if pd.notna(value) and '_YOY' not in col:
                try:
                    num_value = float(value)
                    if num_value != 0:
                        # 转换为亿单位
                        value_yi = convert_to_yi(value)
                        # 获取中文名
                        chinese_name = get_chinese_name(col)
                        if chinese_name is None or chinese_name == col:
                            chinese_name = "-"
                        result_data.append({
                            '科目': col,
                            '中文科目': chinese_name,
                            '数值(亿)': value_yi
                        })
                except:
                    if str(value) not in ['False', 'nan', 'None', '']:
                        chinese_name = get_chinese_name(col)
                        if chinese_name is None or chinese_name == col:
                            chinese_name = "-"
                        result_data.append({
                            '科目': col,
                            '中文科目': chinese_name,
                            '数值(亿)': value
                        })
    
    if not result_data:
        return None
    
    result_df = pd.DataFrame(result_data)

    # 根据报表类型调整显示顺序
    if statement_type in ("profit", "cash_flow"):
        # 利润表常见科目顺序
        profit_order = [
            "TOTAL_OPERATE_INCOME", "OPERATE_INCOME",
            "OPERATE_COST", "OPERATE_TAX_ADD",
            "SALE_EXPENSE", "MANAGE_EXPENSE", "RESEARCH_EXPENSE",
            "FINANCE_EXPENSE",
            "FAIRVALUE_CHANGE_INCOME", "INVEST_INCOME", "OTHER_INCOME",
            "ASSET_DISPOSAL_INCOME", "NONBUSINESS_INCOME", "NONBUSINESS_EXPENSE",
            "TOTAL_PROFIT", "INCOME_TAX",
            "NETPROFIT", "PARENT_NETPROFIT", "DEDUCT_PARENT_NETPROFIT",
            "MINORITY_INTEREST",
            "BASIC_EPS", "DILUTED_EPS",
        ]
        # 现金流量表常见科目顺序（按三大活动）
        cash_order = [
            # 经营活动
            "SALE_SERVICE", "SALES_SERVICES", "RECEIVE_OTHER_OPERATE",
            "OPERATE_INFLOW_BALANCE",
            "BUY_SERVICE", "BUY_SERVICES", "PAY_STAFF_CASH",
            "PAY_ALL_TAX", "PAY_OTHER_OPERATE",
            "OPERATE_NETCASH_OPERATE", "NETCASH_OPERATE", "OPERATE_NET_CASH_FLOW",
            # 投资活动
            "WITHDRAW_INVEST", "RECEIVE_INVEST_INCOME", "DISPOSAL_LONG_ASSET",
            "RECEIVE_OTHER_INVEST",
            "INVEST_INFLOW_BALANCE",
            "INVEST_PAY_CASH", "CONSTRUCT_LONG_ASSET", "PAY_OTHER_INVEST",
            "INVEST_OUTFLOW_BALANCE",
            "NETCASH_INVEST", "INVEST_NET_CASH_FLOW",
            # 筹资活动
            "ACCEPT_INVEST_CASH", "ACCEPT_LOAN_CASH", "ISSUE_BOND",
            "RECEIVE_OTHER_FINANCE",
            "FINANCE_INFLOW_BALANCE",
            "PAY_DEBT_CASH", "ASSIGN_DIVIDEND_PORFIT", "PAY_OTHER_FINANCE",
            "FINANCE_OUTFLOW_BALANCE",
            "FINANCE_NET_CASH_FLOW",
            # 其他
            "RATE_CHANGE_EFFECT",
            "NET_CASH_INCREASE", "BEGIN_CASH", "END_CASH",
        ]
        order_list = profit_order if statement_type == "profit" else cash_order
        # 使用科目（英文列名）排序，未出现在顺序表中的放在后面
        result_df["__order"] = result_df["科目"].apply(
            lambda x: order_list.index(x) if x in order_list else len(order_list) + 1
        )
        result_df = result_df.sort_values(by="__order").drop(columns="__order")

    return result_df

def get_financial_statements(symbol, start_year, end_year):
    """
    获取指定年份范围的三大财务报表数据
    
    参数:
        symbol: 股票代码
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        字典，格式为 {年份: {'balance': DataFrame, 'profit': DataFrame, 'cash_flow': DataFrame}}
    """
    symbol_with_suffix = get_symbol_with_suffix(symbol)
    
    results = {}
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        # 1. 获取资产负债表
        status_text.text("📊 正在获取资产负债表数据...")
        progress_bar.progress(10)
        balance_sheet = ak.stock_balance_sheet_by_report_em(symbol=symbol_with_suffix)
        
        # 2. 获取利润表
        status_text.text("💰 正在获取利润表数据...")
        progress_bar.progress(30)
        profit = ak.stock_profit_sheet_by_report_em(symbol=symbol_with_suffix)
        
        # 3. 获取现金流量表
        status_text.text("💵 正在获取现金流量表数据...")
        progress_bar.progress(50)
        cash_flow = ak.stock_cash_flow_sheet_by_report_em(symbol=symbol_with_suffix)
        
        # 4. 按年份组织数据
        status_text.text("📋 正在按年份组织数据...")
        progress_bar.progress(70)
        
        for year in range(start_year, end_year + 1):
            year_data = {
                'balance': None,
                'profit': None,
                'cash_flow': None
            }
            
            # 格式化资产负债表
            if balance_sheet is not None and not balance_sheet.empty:
                year_data['balance'] = format_statement_data(balance_sheet, year, statement_type="balance")
            
            # 格式化利润表（按常规科目顺序）
            if profit is not None and not profit.empty:
                year_data['profit'] = format_statement_data(profit, year, statement_type="profit")
            
            # 格式化现金流量表（按经营/投资/筹资顺序）
            if cash_flow is not None and not cash_flow.empty:
                year_data['cash_flow'] = format_statement_data(cash_flow, year, statement_type="cash_flow")
            
            # 如果至少有一个报表有数据，就添加到结果中
            has_data = False
            for key, value in year_data.items():
                if value is not None and not value.empty:
                    has_data = True
                    break
            
            if has_data:
                results[year] = year_data
        
        progress_bar.progress(100)
        status_text.text("✅ 数据获取完成！")
        
    except Exception as e:
        st.error(f"❌ 获取数据失败：{str(e)}")
        import traceback
        with st.expander("查看详细错误信息"):
            st.code(traceback.format_exc())
        return None
    
    finally:
        progress_bar.empty()
        status_text.empty()
    
    return results

def create_excel_file(results, symbol, company_name, start_year, end_year):
    """
    创建Excel文件，每年一个sheet，每个sheet包含三大报表
    
    参数:
        results: 数据字典 {年份: {'balance': DataFrame, 'profit': DataFrame, 'cash_flow': DataFrame}}
        symbol: 股票代码
        company_name: 公司名称
        start_year: 起始年份
        end_year: 结束年份
    
    返回:
        Excel文件的字节数据
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for year in sorted(results.keys()):
            year_data = results[year]
            sheet_name = f"{year}年"
            
            # 创建该年份的完整数据框
            year_rows = []
            
            # 1. 资产负债表
            if year_data['balance'] is not None and not year_data['balance'].empty:
                # 添加标题行
                year_rows.append({
                    '科目': '【资产负债表】',
                    '中文科目': '',
                    '数值(亿)': ''
                })
                # 添加空行
                year_rows.append({
                    '科目': '',
                    '中文科目': '',
                    '数值(亿)': ''
                })
                # 添加数据
                for _, row in year_data['balance'].iterrows():
                    year_rows.append({
                        '科目': row['科目'],
                        '中文科目': row['中文科目'],
                        '数值(亿)': row['数值(亿)']
                    })
                # 添加分隔空行
                year_rows.append({
                    '科目': '',
                    '中文科目': '',
                    '数值(亿)': ''
                })
                year_rows.append({
                    '科目': '',
                    '中文科目': '',
                    '数值(亿)': ''
                })
            
            # 2. 利润表
            if year_data['profit'] is not None and not year_data['profit'].empty:
                # 添加标题行
                year_rows.append({
                    '科目': '【利润表】',
                    '中文科目': '',
                    '数值(亿)': ''
                })
                # 添加空行
                year_rows.append({
                    '科目': '',
                    '中文科目': '',
                    '数值(亿)': ''
                })
                # 添加数据
                for _, row in year_data['profit'].iterrows():
                    year_rows.append({
                        '科目': row['科目'],
                        '中文科目': row['中文科目'],
                        '数值(亿)': row['数值(亿)']
                    })
                # 添加分隔空行
                year_rows.append({
                    '科目': '',
                    '中文科目': '',
                    '数值(亿)': ''
                })
                year_rows.append({
                    '科目': '',
                    '中文科目': '',
                    '数值(亿)': ''
                })
            
            # 3. 现金流量表
            if year_data['cash_flow'] is not None and not year_data['cash_flow'].empty:
                # 添加标题行
                year_rows.append({
                    '科目': '【现金流量表】',
                    '中文科目': '',
                    '数值(亿)': ''
                })
                # 添加空行
                year_rows.append({
                    '科目': '',
                    '中文科目': '',
                    '数值(亿)': ''
                })
                # 添加数据
                for _, row in year_data['cash_flow'].iterrows():
                    year_rows.append({
                        '科目': row['科目'],
                        '中文科目': row['中文科目'],
                        '数值(亿)': row['数值(亿)']
                    })
            
            # 创建DataFrame并写入sheet
            if year_rows:
                year_df = pd.DataFrame(year_rows)
                year_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # 设置列宽自适应
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        output.seek(0)
        wb = load_workbook(output)
        
        # 为每个sheet设置列宽
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, values_only=False), start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in col:
                    if cell.value is not None:
                        try:
                            cell_value = str(cell.value)
                            length = 0
                            for char in cell_value:
                                if ord(char) > 127:  # 非ASCII字符（包括中文）
                                    length += 2
                                else:
                                    length += 1
                            if length > max_length:
                                max_length = length
                        except:
                            pass
                
                if max_length > 0:
                    adjusted_width = min(max(max_length + 2, 8), 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                else:
                    ws.column_dimensions[column_letter].width = 10
        
        output.seek(0)
        output.truncate(0)
        wb.save(output)
        wb.close()
    except Exception as e:
        # 如果设置列宽失败，不影响返回结果
        print(f"⚠️ 设置列宽失败: {e}")
    
    output.seek(0)
    return output.getvalue()

# Streamlit界面（仅在直接运行时执行）
if __name__ == "__main__":
    st.set_page_config(
        page_title="财务报表下载工具",
        page_icon="📊",
        layout="wide"
    )

    st.title("📊 财务报表下载工具")
    st.markdown("---")

    # 侧边栏
    with st.sidebar:
        st.header("📝 参数设置")
        
        # 股票代码输入
        symbol = st.text_input(
            "股票代码",
            value="603486",
            help="请输入6位股票代码，如：603486（科沃斯）、600519（贵州茅台）"
        )
        
        # 年份范围
        col1, col2 = st.columns(2)
        with col1:
            start_year = st.number_input(
                "起始年份",
                min_value=2000,
                max_value=datetime.now().year,
                value=2015,
                step=1
            )
        with col2:
            end_year = st.number_input(
                "结束年份",
                min_value=2000,
                max_value=datetime.now().year,
                value=2024,
                step=1
            )
        
        if start_year > end_year:
            st.error("⚠️ 起始年份不能大于结束年份")
            st.stop()
        
        # 开始按钮
        analyze_button = st.button(
            "🚀 开始获取数据",
            type="primary",
            use_container_width=True
        )

    # 主内容区
    if analyze_button:
        if not symbol or len(symbol.replace('.SZ', '').replace('.SH', '')) != 6:
            st.error("❌ 请输入有效的6位股票代码")
            st.stop()
        
        # 获取公司名称
        company_name = get_symbol_name(symbol)
        st.info(f"📌 公司名称：**{company_name}** ({symbol})")
        st.info(f"📅 年份范围：{start_year} - {end_year}")
        
        # 获取数据
        results = get_financial_statements(symbol, start_year, end_year)
        
        if results and len(results) > 0:
            st.success(f"✅ 成功获取 {len(results)} 个年份的数据")
            
            # 显示数据预览
            st.divider()
            st.header("📋 数据预览")
            
            # 为每个年份创建标签页
            tabs = st.tabs([f"{year}年" for year in sorted(results.keys())])
            
            for idx, year in enumerate(sorted(results.keys())):
                with tabs[idx]:
                    year_data = results[year]
                    
                    # 资产负债表
                    if year_data['balance'] is not None and not year_data['balance'].empty:
                        st.subheader("📊 资产负债表")
                        st.dataframe(
                            year_data['balance'],
                            use_container_width=True,
                            hide_index=True
                        )
                    
                    # 利润表
                    if year_data['profit'] is not None and not year_data['profit'].empty:
                        st.subheader("💰 利润表")
                        st.dataframe(
                            year_data['profit'],
                            use_container_width=True,
                            hide_index=True
                        )
                    
                    # 现金流量表
                    if year_data['cash_flow'] is not None and not year_data['cash_flow'].empty:
                        st.subheader("💵 现金流量表")
                        st.dataframe(
                            year_data['cash_flow'],
                            use_container_width=True,
                            hide_index=True
                        )
            
            # 生成Excel文件
            st.divider()
            st.header("📥 下载Excel文件")
            
            excel_data = create_excel_file(results, symbol, company_name, start_year, end_year)
            
            symbol_clean = symbol.replace('.SZ', '').replace('.SH', '')
            filename = f"{company_name}_{symbol_clean}_{start_year}-{end_year}_财务报表.xlsx"
            
            st.download_button(
                label="📥 下载Excel文件",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            st.info(f"💡 文件包含 {len(results)} 个sheet，每个sheet包含该年份的三大财务报表")
            
        else:
            st.warning("⚠️ 未获取到数据，请检查股票代码和年份范围是否正确")

    else:
        # 欢迎页面
        st.markdown("""
        ## 👋 欢迎使用财务报表下载工具
        
        这个工具可以帮助您：
        
        ### ✨ 主要功能
        
        1. **📊 获取历年财务报表**：自动获取指定年份范围的三大财务报表数据
        2. **📋 数据格式化**：将数据转换为易读的格式（每个科目一行）
        3. **📥 Excel导出**：每年一个sheet，每个sheet包含资产负债表、利润表、现金流量表
        4. **💰 单位转换**：自动将数值转换为"亿"单位，便于阅读
        
        ### 🚀 使用步骤
        
        1. 在左侧边栏输入**股票代码**（如：603486）
        2. 选择**起始年份**和**结束年份**
        3. 点击 **"开始获取数据"** 按钮
        4. 查看数据预览
        5. 下载Excel文件
        
        ### 📝 使用示例
        
        - **科沃斯**：603486
        - **贵州茅台**：600519
        - **平安银行**：000001
        - **万科A**：000002
        
        ### ⚠️ 注意事项
        
        - 数据来源于公开数据源，仅供参考
        - 首次获取可能需要较长时间（数据获取）
        - 建议选择合理的年份范围（通常5-10年）
        - 某些股票可能缺少部分年份的数据
        
        ---
        
        **开始使用**：请在左侧边栏设置参数，然后点击"开始获取数据"按钮。
        """)

